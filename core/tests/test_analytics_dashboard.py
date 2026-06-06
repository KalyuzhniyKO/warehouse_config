from decimal import Decimal
from io import BytesIO, StringIO

from openpyxl import load_workbook

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from core.models import Item, Location, Recipient, StockBalance, StockMovement, Unit, Warehouse
from core.services.analytics import get_analytics_summary, get_kpi_delta, get_top_issued_items, get_reconciliation_summary
from core.tests.warehouse_access_utils import grant_warehouse_access


class AnalyticsDashboardTests(TestCase):
    def setUp(self):
        call_command("init_roles", stdout=StringIO())
        self.admin = get_user_model().objects.create_user("adm", password="pw")
        self.admin.groups.add(Group.objects.get(name="Адміністратор складу"))
        self.storekeeper = get_user_model().objects.create_user("keeper", password="pw")
        self.storekeeper.groups.add(Group.objects.get(name="Комірник"))
        self.superuser = get_user_model().objects.create_superuser("root", password="pw", email="r@e.com")
        self.unit = Unit.objects.create(name="шт", symbol="шт")
        self.item = Item.objects.create(name="Кабель", unit=self.unit, internal_code="KB-1")
        self.wh = Warehouse.objects.create(name="WH")
        grant_warehouse_access(self.admin, self.wh, can_delegate=True)
        self.loc = Location.objects.create(name="L1", warehouse=self.wh)
        self.rec = Recipient.objects.create(name="Р1")
        StockBalance.objects.create(item=self.item, location=self.loc, qty=Decimal("5.000"))
        now = timezone.now()
        StockMovement.objects.create(movement_type=StockMovement.MovementType.IN, item=self.item, qty=Decimal("10.000"), destination_location=self.loc, occurred_at=now, document_number="DOC-IN")
        StockMovement.objects.create(movement_type=StockMovement.MovementType.OUT, item=self.item, qty=Decimal("2.000"), source_location=self.loc, recipient=self.rec, department="Цех 1", occurred_at=now, document_number="DOC-OUT")
        StockMovement.objects.create(movement_type=StockMovement.MovementType.OUT, item=self.item, qty=Decimal("1.000"), source_location=self.loc, occurred_at=now - timezone.timedelta(days=2), document_number="DOC-OLD")

    def test_access(self):
        self.assertEqual(self.client.get(reverse("management_analytics")).status_code, 302)
        self.client.force_login(self.storekeeper)
        self.assertEqual(self.client.get(reverse("management_analytics")).status_code, 403)
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse("management_analytics")).status_code, 200)
        self.client.force_login(self.superuser)
        self.assertEqual(self.client.get(reverse("management_analytics")).status_code, 200)

    def test_dashboard_structure_and_export_labels(self):
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics"))
        self.assertContains(r, "Аналітика складу")
        self.assertContains(r, "Огляд рухів, залишків, видачі та якості складських даних.")
        self.assertContains(r, "analytics-kpi-grid")
        self.assertContains(r, "analytics-filter-panel")
        self.assertContains(r, "Експорт Excel")
        self.assertContains(r, "Експорт CSV")
        self.assertNotContains(r, "Експорт PDF")
        self.assertNotContains(r, ">movement_type=out<", html=True)
        self.assertNotContains(r, "data-analytics-daily-chart")
        self.assertNotContains(r, "data-analytics-operation-mix")
        self.assertNotContains(r, "data-analytics-top-items-chart")

    def test_kpi_cards_filter_panel_and_advanced_location_render(self):
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics"))
        for label in ["Операції", "Критичні залишки", "Видано", "Надійшло", "Без руху"]:
            self.assertContains(r, label)
        for label in ["Фільтри", "Період", "Дата від", "Дата до", "Склад", "Тип операції"]:
            self.assertContains(r, label)
        self.assertContains(r, "analyticsAdvancedFilters")
        self.assertContains(r, "Необов'язкова адресна деталізація всередині складу.")

    def test_analytics_visual_blocks_render_existing_data(self):
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics"))
        self.assertContains(r, "Топ товарів по видачі")
        self.assertContains(r, "Структура операцій")
        self.assertContains(r, "Куди видавали")
        self.assertContains(r, "Кому видавали")
        self.assertContains(r, "analytics-visual-list")
        self.assertContains(r, "analytics-visual-bar-fill")
        self.assertContains(r, "Кабель")
        self.assertContains(r, "analytics-visual-value\">3", html=False)
        self.assertContains(r, "width: 100%")
        for label in ["Прихід", "Видача", "Повернення", "Списання", "Переміщення"]:
            self.assertContains(r, label)

    def test_summary_and_top(self):
        summary = get_analytics_summary({})
        self.assertEqual(summary["operations_count"], 3)
        top = get_top_issued_items({})
        self.assertEqual(top[0]["item__name"], "Кабель")

    def test_kpi_delta_helper(self):
        self.assertEqual(get_kpi_delta(10, 5)["trend"], "positive")
        self.assertEqual(get_kpi_delta(0, 0)["trend"], "neutral")
        self.assertEqual(get_kpi_delta(3, 0)["label"], "нові дані")

    def test_export_csv(self):
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics_export_csv"), {"movement_type": StockMovement.MovementType.OUT})
        self.assertEqual(r.status_code, 200)
        self.assertIn("text/csv", r["Content-Type"])
        self.assertIn("analytics", r["Content-Disposition"])
        body = r.content.decode("utf-8")
        self.assertIn("Топ товарів", body)
        self.assertIn("DOC-OUT", body)
        self.assertNotIn("DOC-IN", body)

    def test_export_access_permissions(self):
        csv_url = reverse("management_analytics_export_csv")
        xlsx_url = reverse("management_analytics_export_xlsx")
        self.assertEqual(self.client.get(csv_url).status_code, 302)
        self.assertEqual(self.client.get(xlsx_url).status_code, 302)

        self.client.force_login(self.storekeeper)
        self.assertEqual(self.client.get(csv_url).status_code, 403)
        self.assertEqual(self.client.get(xlsx_url).status_code, 403)

        self.client.force_login(self.superuser)
        self.assertEqual(self.client.get(csv_url).status_code, 200)
        self.assertEqual(self.client.get(xlsx_url).status_code, 200)

    def test_recent_document_and_empty_state(self):
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics"))
        self.assertContains(r, "DOC-OUT")
        self.assertContains(r, "Останні операції")
        self.assertContains(r, "<table", html=False)
        for header in ["Дата", "Документ", "Номенклатура", "Тип", "Кількість", "Склад"]:
            self.assertContains(r, header)
        StockMovement.objects.all().delete()
        r2 = self.client.get(reverse("management_analytics"))
        self.assertContains(r2, "За вибраний період немає складських операцій.")

    def test_detail_pages_and_links_and_quick_filters(self):
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics"), {"date_from": timezone.localdate().isoformat(), "date_to": timezone.localdate().isoformat()})
        self.assertContains(r, reverse("management_analytics_item_detail", args=[self.item.pk]))
        self.assertContains(r, reverse("management_analytics_usage_place_detail", args=["Цех 1"]))
        self.assertContains(r, reverse("management_analytics_recipient_detail", args=[self.rec.pk]))
        self.assertContains(r, "period=today")
        self.assertContains(r, "period=7d")
        self.assertContains(r, "period=30d")
        self.assertContains(r, "period=month")
        self.assertContains(r, "period=prev_month")

    def test_detail_access_permissions(self):
        self.client.force_login(self.storekeeper)
        self.assertEqual(self.client.get(reverse("management_analytics_item_detail", args=[self.item.pk])).status_code, 403)
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse("management_analytics_item_detail", args=[self.item.pk])).status_code, 200)
        self.assertEqual(self.client.get(reverse("management_analytics_usage_place_detail", args=["Цех 1"])).status_code, 200)
        self.assertEqual(self.client.get(reverse("management_analytics_recipient_detail", args=[self.rec.pk])).status_code, 200)


    def test_dashboard_quality_card_marker(self):
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics"))
        self.assertContains(r, "data-analytics-quality-card")
        self.assertContains(r, "Контроль даних")
        self.assertContains(r, "Переглянути деталі")

    def test_data_quality_page_access(self):
        url = reverse("management_analytics_data_quality")
        self.assertEqual(self.client.get(url).status_code, 302)
        self.client.force_login(self.storekeeper)
        self.assertEqual(self.client.get(url).status_code, 403)
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(url).status_code, 200)

    def test_missing_document_is_reported(self):
        StockMovement.objects.create(movement_type=StockMovement.MovementType.OUT, item=self.item, qty=Decimal("1.000"), source_location=self.loc, occurred_at=timezone.now(), document_number="")
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics_data_quality"))
        self.assertContains(r, "Рухи без документа")
        self.assertNotContains(r, "missing_documents")
        self.assertContains(r, "Переглянути в журналі")

    def test_issue_without_recipient_warning(self):
        StockMovement.objects.create(movement_type=StockMovement.MovementType.OUT, item=self.item, qty=Decimal("1.000"), source_location=self.loc, department="Цех 2", occurred_at=timezone.now(), document_number="D")
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics_data_quality"))
        self.assertContains(r, "Видача без отримувача")
        self.assertNotContains(r, "issue_without_recipient")

    def test_negative_stock_warning_is_zero_for_non_negative_balances(self):
        StockBalance.objects.create(item=Item.objects.create(name="Zero", unit=self.unit, internal_code="ZERO"), location=self.loc, qty=Decimal("0.000"))
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics_data_quality"))
        self.assertContains(r, "Негативних залишків")
        self.assertEqual(r.context["data_quality"]["reconciliation"]["negative_stock"], 0)


    def test_data_quality_page_has_human_labels_and_table_headers(self):
        StockMovement.objects.create(movement_type=StockMovement.MovementType.OUT, item=self.item, qty=Decimal("1.000"), source_location=self.loc, department="", occurred_at=timezone.now(), document_number="")
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics_data_quality"))
        self.assertContains(r, "Рухи без документа")
        self.assertContains(r, "Видача без отримувача")
        self.assertContains(r, "Видача без цеху")
        self.assertContains(r, "Некоректна кількість")
        self.assertContains(r, "Переглянути в журналі")
        self.assertContains(r, "Дата")
        self.assertContains(r, "Тип")
        self.assertContains(r, "Товар")
        self.assertContains(r, "Кількість")
        self.assertContains(r, "Документ")
        self.assertContains(r, "Отримувач")
        self.assertContains(r, "data-analytics-quality-detail")
        self.assertContains(r, "analytics-quality-table")
        self.assertNotContains(r, ">missing_documents<", html=True)
        self.assertNotContains(r, ">issue_without_recipient<", html=True)
        self.assertNotContains(r, ">issue_without_usage_place<", html=True)
        self.assertNotContains(r, ">movement_without_item<", html=True)
        self.assertNotContains(r, ">non_positive_qty<", html=True)
        self.assertNotContains(r, ">receive_without_destination<", html=True)

    def test_reconciliation_total_matches_queryset(self):
        summary = get_reconciliation_summary({})
        self.assertEqual(summary["total_movements"], StockMovement.objects.count())
    def test_xlsx_export(self):
        self.client.force_login(self.admin)
        r = self.client.get(reverse("management_analytics_export_xlsx"), {"movement_type": StockMovement.MovementType.OUT})
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheetml", r["Content-Type"])
        wb = load_workbook(filename=BytesIO(r.content))
        for title in ["Summary", "Daily movement", "Operation mix", "Top issued items", "Top usage places", "Top recipients", "Recent movements", "Inactive stock items", "Data quality"]:
            self.assertIn(title, wb.sheetnames)

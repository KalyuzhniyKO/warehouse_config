from datetime import timedelta
from decimal import Decimal
from io import BytesIO, StringIO

from openpyxl import load_workbook
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone, translation

from ..models import (
    InventoryCount,
    Item,
    Location,
    Recipient,
    StockMovement,
    Unit,
    Warehouse,
)
from .i18n_test_utils import compile_test_messages
from .warehouse_access_utils import grant_warehouse_access


class StockOperationAuditTests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        compile_test_messages()

    def setUp(self):
        translation.activate("uk")
        call_command("init_roles", stdout=StringIO())
        User = get_user_model()
        self.admin = User.objects.create_user("operation-audit-admin", password="pw")
        self.admin.groups.add(Group.objects.get(name="Адміністратор складу"))
        self.storekeeper = User.objects.create_user(
            "operation-audit-storekeeper",
            password="pw",
        )
        self.storekeeper.groups.add(Group.objects.get(name="Комірник"))
        self.actor = User.objects.create_user(
            "operation-audit-actor",
            password="pw",
            first_name="Audit",
            last_name="Actor",
        )
        self.canceller = User.objects.create_user(
            "operation-audit-canceller",
            password="pw",
            first_name="Cancel",
            last_name="Manager",
        )
        self.unit = Unit.objects.create(name="Audit piece", symbol="ap")
        self.item = Item.objects.create(
            name="Audit cable",
            internal_code="AUDIT-CABLE",
            unit=self.unit,
        )
        self.other_item = Item.objects.create(name="Other audit item", unit=self.unit)
        self.warehouse = Warehouse.objects.create(name="Audit warehouse")
        self.other_warehouse = Warehouse.objects.create(name="Other audit warehouse")
        grant_warehouse_access(self.admin, [self.warehouse, self.other_warehouse])
        grant_warehouse_access(self.storekeeper, [self.warehouse, self.other_warehouse])
        self.location = Location.objects.create(warehouse=self.warehouse, name="AUDIT-A")
        self.other_location = Location.objects.create(
            warehouse=self.other_warehouse,
            name="AUDIT-B",
        )
        self.recipient = Recipient.objects.create(name="Audit recipient")
        self.inventory = InventoryCount.objects.create(
            number="INV-AUDIT-0001",
            warehouse=self.warehouse,
            status=InventoryCount.Status.COMPLETED,
        )
        self.old_movement = StockMovement.objects.create(
            movement_type=StockMovement.MovementType.IN,
            item=self.other_item,
            qty=Decimal("5.000"),
            destination_warehouse=self.other_warehouse,
            destination_location=self.other_location,
            performed_by=self.actor,
            created_by=self.actor,
            document_number="AUDIT-OLD",
        )
        self.movement = StockMovement.objects.create(
            movement_type=StockMovement.MovementType.OUT,
            item=self.item,
            qty=Decimal("2.000"),
            source_warehouse=self.warehouse,
            source_location=self.location,
            recipient=self.recipient,
            performed_by=self.actor,
            created_by=self.actor,
            document_number="AUDIT-DOC",
            comment="Business audit comment",
            inventory_count=self.inventory,
        )
        self.reversal = StockMovement.objects.create(
            movement_type=StockMovement.MovementType.ADJUSTMENT,
            item=self.item,
            qty=Decimal("2.000"),
            destination_warehouse=self.warehouse,
            destination_location=self.location,
            performed_by=self.canceller,
            created_by=self.canceller,
            reversal_of=self.movement,
            comment="Cancellation movement",
        )
        self.movement.is_cancelled = True
        self.movement.cancelled_by = self.canceller
        self.movement.cancelled_at = timezone.now()
        self.movement.cancellation_reason = "Wrong recipient"
        self.movement.cancellation_movement = self.reversal
        self.movement.save(
            update_fields=[
                "is_cancelled",
                "cancelled_by",
                "cancelled_at",
                "cancellation_reason",
                "cancellation_movement",
                "updated_at",
            ]
        )
        old_created_at = timezone.now() - timedelta(days=10)
        StockMovement.objects.filter(pk=self.old_movement.pk).update(
            created_at=old_created_at
        )
        self.old_movement.refresh_from_db()
        self.url = reverse("stock_operation_audit")
        self.export_url = reverse("stock_operation_audit_export_xlsx")

    def audit_response(self, params=None, user=None):
        self.client.force_login(user or self.admin)
        return self.client.get(self.url, params or {})

    def movement_ids(self, response):
        return [movement.pk for movement in response.context["movements"]]

    def export_workbook(self, params=None, user=None):
        self.client.force_login(user or self.admin)
        response = self.client.get(self.export_url, params or {})
        return response, load_workbook(BytesIO(response.content), data_only=True)

    def test_management_user_can_open_operation_audit(self):
        response = self.audit_response()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Аудит операцій")
        self.assertNotContains(response, "filter-panel")
        self.assertContains(response, "operation-audit-filter-form")
        self.assertContains(response, "table-filter-heading")
        self.assertContains(response, "table-filter-toggle")
        self.assertContains(response, 'name="date_from"')
        self.assertContains(response, 'name="movement_type"')
        self.assertContains(response, 'name="q"')
        self.assertContains(response, 'name="quantity_from"')
        self.assertContains(response, 'name="quantity_to"')
        self.assertContains(response, 'name="warehouse"')
        self.assertContains(response, 'name="location"')
        self.assertContains(response, 'name="recipient"')
        self.assertContains(response, 'name="document"')
        self.assertContains(response, 'name="user"')
        self.assertContains(response, 'name="cancelled"')
        self.assertContains(response, 'name="inventory_related"')

    def test_non_management_user_cannot_open_operation_audit(self):
        response = self.audit_response(user=self.storekeeper)

        self.assertEqual(response.status_code, 403)

    def test_page_shows_business_fields_and_linked_records(self):
        response = self.audit_response()

        self.assertContains(response, "Audit cable")
        self.assertEqual(response.context["movements"][1].qty, Decimal("2.000"))
        self.assertContains(response, "Audit warehouse")
        self.assertContains(response, "Audit recipient")
        self.assertContains(response, "AUDIT-DOC")
        self.assertContains(response, "Business audit comment")
        self.assertContains(response, "Audit Actor")
        self.assertContains(response, "Cancel Manager")
        self.assertContains(response, "Wrong recipient")
        self.assertContains(response, "Анулювано")
        self.assertContains(response, "Пов'язано з інвентаризацією")
        self.assertContains(response, reverse("inventory_detail", args=[self.inventory.pk]))
        self.assertContains(response, reverse("stock_movement_print", args=[self.reversal.pk]))

    def test_filters_by_date_operation_type_warehouse_and_cancelled(self):
        today = timezone.localdate().isoformat()

        response = self.audit_response(
            {
                "date_from": today,
                "date_to": today,
                "movement_type": StockMovement.MovementType.OUT,
                "warehouse": self.warehouse.pk,
                "cancelled": "yes",
            }
        )

        self.assertEqual(self.movement_ids(response), [self.movement.pk])

    def test_cancelled_no_filter_excludes_cancelled_movement(self):
        response = self.audit_response({"cancelled": "no"})

        self.assertNotIn(self.movement.pk, self.movement_ids(response))
        self.assertIn(self.old_movement.pk, self.movement_ids(response))

    def test_filters_by_item_recipient_user_and_inventory_relation(self):
        response = self.audit_response(
            {
                "q": "AUDIT-CABLE",
                "recipient": self.recipient.pk,
                "user": self.actor.pk,
                "inventory_related": "yes",
            }
        )

        self.assertEqual(self.movement_ids(response), [self.movement.pk])

    def test_filters_by_quantity_location_and_document_text(self):
        response = self.audit_response(
            {
                "quantity_from": "1.500",
                "quantity_to": "2.500",
                "location": self.location.pk,
                "document": "Business audit",
            }
        )

        self.assertEqual(self.movement_ids(response), [self.movement.pk])

    def test_newest_movements_are_listed_first(self):
        response = self.audit_response()

        self.assertEqual(self.movement_ids(response)[0], self.reversal.pk)
        self.assertEqual(self.movement_ids(response)[-1], self.old_movement.pk)

    def test_localized_operation_audit_labels(self):
        expectations = {
            "uk": "Аудит операцій",
            "en": "Operation audit",
            "ru": "Аудит операций",
        }

        for language_code, label in expectations.items():
            self.client.force_login(self.admin)
            response = self.client.get(f"/{language_code}{self.url[3:]}")
            with self.subTest(language_code=language_code):
                self.assertContains(response, label)

    def test_export_requires_management_permission(self):
        self.client.force_login(self.storekeeper)

        response = self.client.get(self.export_url)

        self.assertEqual(response.status_code, 403)

    def test_export_has_business_sheets_headers_and_filename(self):
        response, workbook = self.export_workbook()

        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
        self.assertRegex(
            response["Content-Disposition"],
            r'attachment; filename="stock-operation-audit-\d{8}-\d{4}\.xlsx"',
        )
        self.assertEqual(workbook.sheetnames, ["Summary", "Audit report"])
        report = workbook["Audit report"]
        self.assertEqual(
            [cell.value for cell in report[1]],
            [
                "Дата й час операції",
                "Дата й час створення",
                "Тип операції",
                "Статус",
                "Код товару",
                "Назва товару",
                "Кількість",
                "Одиниця",
                "Склад-відправник",
                "Склад-отримувач",
                "Локація-відправник",
                "Локація-отримувач",
                "Отримувач",
                "Документ",
                "Коментар / причина",
                "Створив",
                "Анулювано",
                "Анулював",
                "Час анулювання",
                "Рух анулювання",
                "Зворотний рух",
                "Інвентаризація",
            ],
        )
        self.assertEqual(report.max_row - 1, StockMovement.objects.count())
        self.assertEqual(report.freeze_panes, "A2")
        self.assertEqual(report.auto_filter.ref, report.dimensions)
        summary_values = [cell.value for cell in workbook["Summary"]["A"]]
        self.assertIn("Звіт аудиту складських операцій", summary_values)
        self.assertIn("Усього рухів", summary_values)

    def test_export_respects_operation_type_filter(self):
        _, workbook = self.export_workbook(
            {"movement_type": StockMovement.MovementType.OUT}
        )
        report = workbook["Audit report"]

        self.assertEqual(report.max_row, 2)
        self.assertEqual(report.cell(row=2, column=5).value, "AUDIT-CABLE")
        self.assertEqual(report.cell(row=2, column=14).value, "AUDIT-DOC")

    def test_export_respects_warehouse_and_cancelled_filters(self):
        _, warehouse_workbook = self.export_workbook(
            {"warehouse": self.other_warehouse.pk}
        )
        _, cancelled_workbook = self.export_workbook({"cancelled": "yes"})

        warehouse_report = warehouse_workbook["Audit report"]
        cancelled_report = cancelled_workbook["Audit report"]
        self.assertEqual(warehouse_report.max_row, 2)
        self.assertEqual(
            warehouse_report.cell(row=2, column=10).value,
            "Other audit warehouse",
        )
        self.assertEqual(cancelled_report.max_row, 2)
        self.assertEqual(cancelled_report.cell(row=2, column=17).value, "Так")
        self.assertEqual(cancelled_report.cell(row=2, column=22).value, "INV-AUDIT-0001")

    def test_export_link_preserves_page_filters(self):
        response = self.audit_response(
            {"movement_type": StockMovement.MovementType.OUT, "cancelled": "yes"}
        )

        self.assertContains(response, self.export_url)
        self.assertContains(response, "movement_type=out")
        self.assertContains(response, "cancelled=yes")

    def test_export_headers_are_localized(self):
        expectations = {
            "en": "Operation date/time",
            "ru": "Дата и время операции",
            "it": "Data/ora operazione",
            "pl": "Data i czas operacji",
        }
        self.client.force_login(self.admin)

        for language_code, expected_header in expectations.items():
            response = self.client.get(
                f"/{language_code}{self.export_url[3:]}",
                {"movement_type": StockMovement.MovementType.OUT},
            )
            workbook = load_workbook(BytesIO(response.content), data_only=True)
            with self.subTest(language_code=language_code):
                self.assertEqual(
                    workbook["Audit report"].cell(row=1, column=1).value,
                    expected_header,
                )

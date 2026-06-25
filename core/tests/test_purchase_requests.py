from io import BytesIO, StringIO
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.test import TestCase
from django.urls import reverse
from django.utils import translation
from django.utils import timezone

from core.forms import PurchaseRequestForm
from core.models import Item, PurchaseRequest, StockBalance, StockMovement, Unit, Warehouse
from core.services.purchase_requests import (
    archive_purchase_request,
    restore_purchase_request,
)
from core.tests.i18n_test_utils import compile_test_messages
from core.tests.warehouse_access_utils import grant_warehouse_access


class PurchaseRequestTests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        compile_test_messages()

    def setUp(self):
        from django.core.management import call_command

        call_command("init_roles", stdout=StringIO())
        User = get_user_model()
        self.admin = User.objects.create_user("purchase-admin", password="pw")
        self.admin.groups.add(Group.objects.get(name="Адміністратор складу"))
        self.requester = User.objects.create_user("purchase-requester", password="pw")
        self.other_requester = User.objects.create_user("purchase-other", password="pw")
        self.no_access_user = User.objects.create_user("purchase-no-access", password="pw")
        self.warehouse = Warehouse.objects.create(name="Purchase warehouse")
        self.unit = Unit.objects.create(name="Pair", symbol="pairs")
        self.item = Item.objects.create(name="Safety gloves", unit=self.unit)
        grant_warehouse_access(self.admin, self.warehouse, can_delegate=True)
        grant_warehouse_access(self.requester, self.warehouse)
        grant_warehouse_access(self.other_requester, self.warehouse)

    def request_data(self, **overrides):
        data = {
            "request_date": "2026-06-15",
            "requested_item": "Safety gloves",
            "title": "Safety gloves",
            "need_description": "Planned purchase",
            "requested_qty": "12.000",
            "unit": "pairs",
            "unit_price_uah": "55.50",
            "order_type": PurchaseRequest.OrderType.PLANNED,
            "product_url": "",
            "comment": "Needed next month",
        }
        data.update(overrides)
        return data

    def create_request(self, user=None, status=PurchaseRequest.Status.DRAFT, **overrides):
        data = self.request_data(**overrides)
        data.pop("requested_item", None)
        purchase_request = PurchaseRequest.objects.create(
            requested_by=user or self.requester,
            status=status,
            **data,
        )
        purchase_request.refresh_from_db()
        return purchase_request

    def login(self, user):
        self.client.force_login(user)

    def grant_permission(self, user, codename):
        user.user_permissions.add(Permission.objects.get(codename=codename))

    def post_action(self, purchase_request, action, user=None):
        self.login(user or self.admin)
        return self.client.post(reverse(f"purchase_request_{action}", args=[purchase_request.pk]))

    def load_purchase_request_export(self, params=None, user=None):
        from openpyxl import load_workbook

        self.login(user or self.admin)
        response = self.client.get(reverse("purchase_request_export_xlsx"), params or {})
        workbook = load_workbook(BytesIO(response.content))
        return response, workbook.active

    def load_purchase_request_archive_export(self, params=None, user=None):
        from openpyxl import load_workbook

        self.login(user or self.admin)
        response = self.client.get(
            reverse("purchase_request_archive_export_xlsx"), params or {}
        )
        workbook = load_workbook(BytesIO(response.content))
        return response, workbook.active

    def test_user_with_warehouse_access_can_create_purchase_request(self):
        self.login(self.requester)

        response = self.client.post(reverse("purchase_request_create"), self.request_data())

        purchase_request = PurchaseRequest.objects.get()
        self.assertRedirects(
            response, reverse("purchase_request_detail", args=[purchase_request.pk])
        )
        self.assertEqual(purchase_request.requested_by, self.requester)
        self.assertEqual(purchase_request.status, PurchaseRequest.Status.PENDING_APPROVAL)
        self.assertEqual(
            purchase_request.approval_status, PurchaseRequest.ApprovalStatus.PENDING
        )
        self.assertEqual(
            purchase_request.payment_status,
            PurchaseRequest.PaymentStatus.INVOICE_NOT_RECEIVED,
        )
        self.assertEqual(
            purchase_request.delivery_status, PurchaseRequest.DeliveryStatus.NOT_SHIPPED
        )
        self.assertEqual(purchase_request.request_date, timezone.localdate())
        self.assertEqual(purchase_request.item, self.item)

    def test_warehouse_admin_can_create_without_assigned_warehouse(self):
        self.admin.warehouse_accesses.all().delete()
        self.login(self.admin)

        response = self.client.post(reverse("purchase_request_create"), self.request_data())

        self.assertEqual(response.status_code, 302)
        self.assertEqual(PurchaseRequest.objects.get().requested_by, self.admin)

    def test_product_url_is_optional(self):
        self.login(self.requester)

        response = self.client.post(
            reverse("purchase_request_create"), self.request_data(product_url="")
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(PurchaseRequest.objects.get().product_url, "")

    def test_create_form_is_minimal_and_hides_tracking_and_audit_fields(self):
        self.login(self.requester)

        response = self.client.get(reverse("purchase_request_create"))

        fields = list(response.context["form"].fields)
        self.assertEqual(
            fields,
            [
                "requested_item",
                "requested_qty",
                "unit",
                "need_description",
                "product_url",
                "order_type",
            ],
        )
        for field in [
            "request_date",
            "title",
            "unit_price_uah",
            "approval_status",
            "payment_status",
            "delivery_status",
            "requested_by",
            "approved_by",
            "approved_at",
            "rejected_by",
            "rejected_at",
            "received_qty",
        ]:
            self.assertNotIn(field, response.context["form"].fields)
            self.assertNotContains(response, f'name="{field}"')
        self.assertNotContains(response, "<table", html=False)

    def test_create_form_uses_compact_layout_fields(self):
        self.login(self.requester)

        response = self.client.get(reverse("purchase_request_create"))

        self.assertContains(response, "purchase-create-form")
        self.assertContains(response, "purchase-create-grid")
        self.assertContains(response, "purchase-create-field--item")
        self.assertContains(response, "purchase-create-field--qty")
        self.assertContains(response, "purchase-create-field--unit")
        self.assertContains(response, "purchase-create-field--description")
        self.assertContains(response, "purchase-create-field--compact-text")
        self.assertContains(response, "purchase-create-field--url")
        self.assertContains(response, "purchase-create-field--wide-url")
        self.assertContains(response, "purchase-create-actions")
        self.assertContains(response, "purchase-create-submit")
        self.assertContains(response, "purchase-create-submit--large")
        self.assertContains(response, 'list="purchase-item-options"')
        self.assertContains(response, '<input type="number" name="requested_qty"', html=False)
        self.assertContains(response, '<select name="unit"', html=False)
        self.assertContains(response, '<option value="pairs"', html=False)
        self.assertContains(response, '<input type="text" name="need_description"', html=False)
        self.assertNotContains(response, '<textarea name="need_description"', html=False)
        self.assertNotContains(response, "purchase-unit-options")

    def test_create_form_orders_units_by_active_item_usage(self):
        meter = Unit.objects.create(name="Meter", symbol="m")
        pack = Unit.objects.create(name="Pack", symbol="pack")
        for index in range(2):
            Item.objects.create(name=f"Meter item {index}", unit=meter)
        Item.objects.create(name="Archived pack item", unit=pack, is_active=False)

        form = PurchaseRequestForm()

        choices = [value for value, label in form.fields["unit"].choices]
        self.assertLess(choices.index("m"), choices.index("pairs"))
        self.assertLess(choices.index("pairs"), choices.index("pack"))

    def test_new_requested_item_text_does_not_create_item(self):
        self.login(self.requester)
        item_count = Item.objects.count()

        response = self.client.post(
            reverse("purchase_request_create"),
            self.request_data(requested_item="Custom bearing", title="Ignored title"),
        )

        purchase_request = PurchaseRequest.objects.get(title="Custom bearing")
        self.assertEqual(response.status_code, 302)
        self.assertIsNone(purchase_request.item)
        self.assertEqual(Item.objects.count(), item_count)

    def test_unit_price_is_optional_and_total_is_empty_without_it(self):
        self.login(self.requester)

        response = self.client.post(
            reverse("purchase_request_create"), self.request_data(unit_price_uah="")
        )

        purchase_request = PurchaseRequest.objects.get()
        self.assertEqual(response.status_code, 302)
        self.assertIsNone(purchase_request.unit_price_uah)
        self.assertIsNone(purchase_request.total_price_uah)

    def test_total_price_is_quantity_times_unit_price(self):
        purchase_request = self.create_request(
            requested_qty="3.000", unit_price_uah="12.50"
        )

        self.assertEqual(purchase_request.total_price_uah, Decimal("37.50"))

    def test_archive_metadata_marks_request_archived_and_restore_clears_it(self):
        purchase_request = self.create_request()

        self.assertIsNone(purchase_request.archived_at)
        self.assertFalse(purchase_request.is_archived)

        archive_purchase_request(
            purchase_request, archived_by=self.admin, reason="Done manually"
        )
        purchase_request.refresh_from_db()

        self.assertTrue(purchase_request.is_archived)
        self.assertEqual(purchase_request.archived_by, self.admin)
        self.assertEqual(purchase_request.archive_reason, "Done manually")

        restore_purchase_request(purchase_request)
        purchase_request.refresh_from_db()

        self.assertFalse(purchase_request.is_archived)
        self.assertIsNone(purchase_request.archived_by)
        self.assertEqual(purchase_request.archive_reason, "")

    def test_requested_qty_is_required_and_positive(self):
        self.login(self.requester)
        url = reverse("purchase_request_create")

        missing = self.client.post(url, self.request_data(requested_qty=""))
        non_positive = self.client.post(url, self.request_data(requested_qty="0"))

        self.assertEqual(missing.status_code, 200)
        self.assertEqual(non_positive.status_code, 200)
        self.assertTrue(missing.context["form"].errors["requested_qty"])
        self.assertTrue(non_positive.context["form"].errors["requested_qty"])
        self.assertEqual(PurchaseRequest.objects.count(), 0)

    def test_admin_list_contains_all_requests_and_regular_list_only_own(self):
        own = self.create_request()
        other = self.create_request(user=self.other_requester, title="Other request")

        self.login(self.requester)
        regular_response = self.client.get(reverse("purchase_request_list"))
        self.assertContains(regular_response, own.title)
        self.assertNotContains(regular_response, other.title)

        self.login(self.admin)
        admin_response = self.client.get(reverse("purchase_request_list"))
        self.assertContains(admin_response, own.title)
        self.assertContains(admin_response, other.title)

    def test_list_shows_requester_name(self):
        self.requester.first_name = "Ivan"
        self.requester.last_name = "Petrenko"
        self.requester.save(update_fields=["first_name", "last_name"])
        self.create_request()
        self.login(self.admin)

        response = self.client.get(reverse("purchase_request_list"))

        self.assertContains(response, "Ivan Petrenko")

    def test_list_filters_are_compact_in_table_header(self):
        self.login(self.admin)

        response = self.client.get(
            reverse("purchase_request_list"),
            {
                "q": "gloves",
                "order_type": PurchaseRequest.OrderType.PLANNED,
                "approval_status": PurchaseRequest.ApprovalStatus.PENDING,
                "payment_status": PurchaseRequest.PaymentStatus.INVOICE_NOT_RECEIVED,
                "delivery_status": PurchaseRequest.DeliveryStatus.NOT_SHIPPED,
                "requested_by": self.requester.pk,
                "date_from": "2026-06-01",
                "date_to": "2026-06-30",
            },
        )

        self.assertContains(response, "purchase-filter-panel")
        self.assertContains(response, "purchase-filter-panel--active")
        self.assertContains(response, "purchase-filter-active-count")
        self.assertContains(response, "активн")
        self.assertContains(response, "purchase-filter-grid")
        self.assertContains(response, "purchase-list-tabs")
        self.assertContains(response, "purchase-list-tab active")
        self.assertNotContains(response, "purchase-table-filter-row")
        self.assertNotContains(response, "purchase-filter-toggle")
        self.assertNotContains(response, "purchase-table-filter-actions")
        self.assertContains(response, 'name="date_from"')
        self.assertContains(response, 'name="date_to"')
        self.assertContains(response, 'name="q"')
        self.assertContains(response, 'value="gloves"')
        self.assertContains(response, 'name="order_type"')
        self.assertContains(response, 'name="approval_status"')
        self.assertContains(response, 'name="payment_status"')
        self.assertContains(response, 'name="delivery_status"')
        self.assertContains(response, 'name="requested_by"')
        self.assertContains(response, 'id="purchase-filter-form"')

    def test_purchase_list_table_is_compact_for_small_screens(self):
        purchase_request = self.create_request(
            title="Compact purchase row",
            need_description="Compact need description",
            requested_qty=Decimal("10.000"),
            unit="pairs",
        )
        self.login(self.admin)

        response = self.client.get(reverse("purchase_request_list"))
        detail_url = reverse("purchase_request_detail", args=[purchase_request.pk])

        self.assertNotContains(response, "<th>Дії</th>", html=False)
        self.assertNotContains(response, "<th>Actions</th>", html=False)
        self.assertNotContains(response, "Перегляд")
        self.assertNotContains(response, "<th>Опис потреби</th>", html=False)
        self.assertNotContains(response, "<th>Кількість</th>", html=False)
        self.assertNotContains(response, "<th>Одиниці вимірювання</th>", html=False)
        self.assertNotContains(
            response, "<th>Вартість за одиницю (грн)</th>", html=False
        )
        self.assertNotContains(response, "<th>Сума (грн)</th>", html=False)
        self.assertNotContains(response, "<th>Посилання на товар</th>", html=False)
        self.assertContains(response, '<th class="purchase-col-status">Погодження</th>', html=False)
        self.assertContains(response, '<th class="purchase-col-status purchase-col-payment-status">Оплата</th>', html=False)
        self.assertContains(response, '<th class="purchase-col-status purchase-col-delivery-status">Доставка</th>', html=False)
        self.assertContains(response, "<th class=\"purchase-col-qty\">К-сть / Од.</th>", html=False)
        self.assertContains(response, "purchase-status-menu")
        self.assertContains(response, "purchase-status-select")
        self.assertNotContains(response, "purchase-request-status-select")
        self.assertNotContains(response, "purchase-request-status-submit")
        self.assertNotContains(response, "min-width: 1280px")
        self.assertContains(response, f'href="{detail_url}"')
        self.assertContains(response, "Compact purchase row")
        self.assertContains(response, "Compact need description")
        self.assertContains(response, "10,000 pairs")

    def test_active_list_hides_archived_requests_and_archive_list_hides_active(self):
        active = self.create_request(title="Active request")
        archived = self.create_request(title="Archived request")
        archive_purchase_request(archived, archived_by=self.admin, reason="Finished")
        self.login(self.admin)

        active_response = self.client.get(reverse("purchase_request_list"))
        archive_response = self.client.get(reverse("purchase_request_archive"))

        self.assertContains(active_response, active.title)
        self.assertNotContains(active_response, archived.title)
        self.assertContains(archive_response, "Архів заявок на закупівлю")
        self.assertContains(archive_response, archived.title)
        self.assertNotContains(archive_response, active.title)
        self.assertContains(archive_response, "Finished")

    def test_archive_list_uses_compact_normal_table_rows(self):
        archived = self.create_request(title="Compact archived request")
        archive_purchase_request(archived, archived_by=self.admin, reason="Finished")
        self.login(self.admin)

        response = self.client.get(reverse("purchase_request_archive"))
        html = response.content.decode()
        header_end = html.index("</thead>")
        row_start = html.index("purchase-request-row", header_end)
        title_start = html.index(archived.title, row_start)
        between_header_and_row = html[header_end:row_start]
        row_tag = html[row_start - 80 : html.index(">", row_start) + 1]

        self.assertContains(response, "purchase-request-table")
        self.assertContains(response, "purchase-request-table-body")
        self.assertContains(response, "purchase-request-row")
        self.assertContains(response, ".purchase-request-table > tbody")
        self.assertContains(response, "display: table-row-group")
        self.assertContains(response, "display: table-row")
        self.assertContains(response, "height: auto")
        self.assertLess(row_start, title_start)
        self.assertNotIn("purchase-request-empty-row", between_header_and_row)
        self.assertNotIn("h-100", row_tag)
        self.assertNotIn("min-vh", row_tag)
        self.assertNotIn("align-items", row_tag)

    def test_active_and_archive_filters_work_independently(self):
        active_match = self.create_request(title="Need filter active")
        active_other = self.create_request(title="Other active")
        archived_match = self.create_request(
            title="Need filter archive",
            payment_status=PurchaseRequest.PaymentStatus.PAID,
        )
        archived_other = self.create_request(title="Other archive")
        archive_purchase_request(archived_match, archived_by=self.admin, reason="Done")
        archive_purchase_request(archived_other, archived_by=self.admin, reason="Done")
        self.login(self.admin)

        active_response = self.client.get(
            reverse("purchase_request_list"), {"q": "filter"}
        )
        archive_response = self.client.get(
            reverse("purchase_request_archive"),
            {"q": "filter", "payment_status": PurchaseRequest.PaymentStatus.PAID},
        )

        self.assertContains(active_response, active_match.title)
        self.assertNotContains(active_response, active_other.title)
        self.assertNotContains(active_response, archived_match.title)
        self.assertContains(archive_response, archived_match.title)
        self.assertNotContains(archive_response, archived_other.title)
        self.assertNotContains(archive_response, active_match.title)

    def test_purchase_list_reset_filters_link_clears_query(self):
        self.login(self.admin)

        response = self.client.get(
            reverse("purchase_request_list"),
            {"q": "gloves", "order_type": PurchaseRequest.OrderType.PLANNED},
        )

        self.assertContains(
            response,
            f'href="{reverse("purchase_request_list")}"',
            html=False,
        )

    def test_list_filters_by_tracking_fields_requester_date_and_search(self):
        matching = self.create_request(
            status=PurchaseRequest.Status.APPROVED,
            order_type=PurchaseRequest.OrderType.URGENT,
            approval_status=PurchaseRequest.ApprovalStatus.APPROVED,
            payment_status=PurchaseRequest.PaymentStatus.PAID,
            delivery_status=PurchaseRequest.DeliveryStatus.IN_TRANSIT,
            title="Matching purchase",
            product_url="https://example.com/target-product",
        )
        self.create_request(user=self.other_requester, title="Unrelated")
        self.login(self.admin)

        response = self.client.get(
            reverse("purchase_request_list"),
            {
                "order_type": PurchaseRequest.OrderType.URGENT,
                "approval_status": PurchaseRequest.ApprovalStatus.APPROVED,
                "payment_status": PurchaseRequest.PaymentStatus.PAID,
                "delivery_status": PurchaseRequest.DeliveryStatus.IN_TRANSIT,
                "requested_by": self.requester.pk,
                "date_from": matching.request_date.isoformat(),
                "date_to": matching.request_date.isoformat(),
                "q": "target-product",
            },
        )

        self.assertContains(response, matching.title)
        self.assertNotContains(response, "Unrelated")

    def test_admin_can_update_tracking_status_from_action_without_stock_effects(self):
        purchase_request = self.create_request()
        self.login(self.admin)
        movement_count = StockMovement.objects.count()
        balance_count = StockBalance.objects.count()

        response = self.client.post(
            reverse("purchase_request_tracking_status", args=[purchase_request.pk]),
            {
                "payment_status": PurchaseRequest.PaymentStatus.PAID,
                "delivery_status": PurchaseRequest.DeliveryStatus.IN_TRANSIT,
            },
        )

        purchase_request.refresh_from_db()
        self.assertRedirects(
            response, reverse("purchase_request_detail", args=[purchase_request.pk])
        )
        self.assertEqual(purchase_request.payment_status, PurchaseRequest.PaymentStatus.PAID)
        self.assertEqual(
            purchase_request.delivery_status,
            PurchaseRequest.DeliveryStatus.IN_TRANSIT,
        )
        self.assertEqual(StockMovement.objects.count(), movement_count)
        self.assertEqual(StockBalance.objects.count(), balance_count)

    def test_regular_user_cannot_update_tracking_status_action(self):
        purchase_request = self.create_request()
        self.login(self.requester)

        response = self.client.post(
            reverse("purchase_request_tracking_status", args=[purchase_request.pk]),
            {"payment_status": PurchaseRequest.PaymentStatus.PAID},
        )

        purchase_request.refresh_from_db()
        self.assertEqual(response.status_code, 403)
        self.assertEqual(
            purchase_request.payment_status,
            PurchaseRequest.PaymentStatus.INVOICE_NOT_RECEIVED,
        )

    def test_list_and_detail_show_tracking_status_controls_for_admin_only(self):
        purchase_request = self.create_request()

        self.login(self.admin)
        list_response = self.client.get(reverse("purchase_request_list"))
        detail_response = self.client.get(
            reverse("purchase_request_detail", args=[purchase_request.pk])
        )
        self.assertContains(list_response, reverse("purchase_request_tracking_status", args=[purchase_request.pk]))
        self.assertContains(detail_response, "detail-payment-status")
        self.assertContains(detail_response, "detail-delivery-status")

        self.login(self.requester)
        list_response = self.client.get(reverse("purchase_request_list"))
        detail_response = self.client.get(
            reverse("purchase_request_detail", args=[purchase_request.pk])
        )
        self.assertNotContains(list_response, reverse("purchase_request_tracking_status", args=[purchase_request.pk]))
        self.assertNotContains(detail_response, "detail-payment-status")
        self.assertNotContains(detail_response, "detail-delivery-status")

    def test_purchase_request_xlsx_export_respects_filters(self):
        matching = self.create_request(
            title="Export target",
            need_description="Need export",
            product_url="https://example.com/export-target",
            payment_status=PurchaseRequest.PaymentStatus.PAID,
            delivery_status=PurchaseRequest.DeliveryStatus.IN_TRANSIT,
            approved_by=self.admin,
            approved_at=timezone.now(),
            rejected_by=self.other_requester,
            rejected_at=timezone.now(),
            rejection_comment="Not now",
        )
        self.create_request(title="Hidden request")

        response, sheet = self.load_purchase_request_export({"q": "export-target"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertRegex(
            response["Content-Disposition"],
            r'attachment; filename="purchase_requests_active_\d{4}-\d{2}-\d{2}\.xlsx"',
        )
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            headers,
            [
                "Дата створення",
                "Назва товару",
                "Опис потреби",
                "Посилання на товар",
                "Кількість",
                "Одиниця виміру",
                "Вартість за одиницю, грн",
                "Сума, грн",
                "Тип замовлення",
                "Статус погодження",
                "Статус оплати",
                "Статус доставки",
                "Заявник",
                "Ким погоджено",
                "Дата погодження",
                "Ким відхилено",
                "Дата відхилення",
                "Коментар відхилення",
            ],
        )
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, sheet.dimensions)
        self.assertTrue(all(cell.font.bold for cell in sheet[1]))
        self.assertEqual(sheet["G2"].number_format, "#,##0.00")
        self.assertEqual(sheet["H2"].number_format, "#,##0.00")
        self.assertTrue(sheet["B2"].alignment.wrap_text)
        self.assertTrue(sheet["C2"].alignment.wrap_text)
        self.assertTrue(sheet["D2"].alignment.wrap_text)
        self.assertTrue(sheet["R2"].alignment.wrap_text)
        self.assertNotEqual(sheet["J2"].fill.fgColor.rgb, "00000000")
        self.assertNotEqual(sheet["K2"].fill.fgColor.rgb, "00000000")
        titles = [row[1].value for row in sheet.iter_rows(min_row=2)]
        self.assertEqual(titles, [matching.title])
        exported_row = next(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(exported_row[3], "https://example.com/export-target")
        self.assertEqual(exported_row[15], self.other_requester.username)
        self.assertEqual(exported_row[17], "Not now")

    def test_purchase_request_xlsx_export_button_preserves_filters(self):
        self.login(self.admin)

        response = self.client.get(
            reverse("purchase_request_list"),
            {"q": "gloves", "payment_status": PurchaseRequest.PaymentStatus.PAID},
        )

        export_url = (
            reverse("purchase_request_export_xlsx")
            + f"?q=gloves&amp;payment_status={PurchaseRequest.PaymentStatus.PAID}"
        )
        self.assertContains(response, "Експорт Excel")
        self.assertContains(response, export_url, html=False)

    def test_purchase_request_xlsx_export_without_filters_includes_visible_requests(self):
        own = self.create_request(title="Visible own export")
        other = self.create_request(user=self.other_requester, title="Visible other export")
        archived = self.create_request(title="Archived hidden export")
        archive_purchase_request(archived, archived_by=self.admin, reason="Finished")

        _response, sheet = self.load_purchase_request_export()

        titles = {row[1].value for row in sheet.iter_rows(min_row=2)}
        self.assertIn(own.title, titles)
        self.assertIn(other.title, titles)
        self.assertNotIn(archived.title, titles)

    def test_purchase_request_archive_xlsx_export_includes_archive_fields_only(self):
        active = self.create_request(title="Active hidden export")
        archived = self.create_request(
            title="Archived export target",
            payment_status=PurchaseRequest.PaymentStatus.PAID,
        )
        archive_purchase_request(archived, archived_by=self.admin, reason="Finished")

        response, sheet = self.load_purchase_request_archive_export(
            {"payment_status": PurchaseRequest.PaymentStatus.PAID}
        )

        self.assertEqual(response.status_code, 200)
        self.assertRegex(
            response["Content-Disposition"],
            r'attachment; filename="purchase_requests_archive_\d{4}-\d{2}-\d{2}\.xlsx"',
        )
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers[-3:], ["Дата архівування", "Архівував", "Причина архівування"])
        titles = [row[1].value for row in sheet.iter_rows(min_row=2)]
        self.assertEqual(titles, [archived.title])
        self.assertNotIn(active.title, titles)
        row = next(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(row[19], self.admin.username)
        self.assertEqual(row[20], "Finished")

    def test_purchase_request_xlsx_export_filters_by_statuses_and_date(self):
        matching = self.create_request(
            title="Export status target",
            request_date="2026-06-10",
            approval_status=PurchaseRequest.ApprovalStatus.APPROVED,
            payment_status=PurchaseRequest.PaymentStatus.PAID,
            delivery_status=PurchaseRequest.DeliveryStatus.DELIVERED,
        )
        self.create_request(
            title="Export status hidden",
            request_date="2026-06-20",
            approval_status=PurchaseRequest.ApprovalStatus.PENDING,
            payment_status=PurchaseRequest.PaymentStatus.INVOICE_NOT_RECEIVED,
            delivery_status=PurchaseRequest.DeliveryStatus.NOT_SHIPPED,
        )

        _response, sheet = self.load_purchase_request_export(
            {
                "approval_status": PurchaseRequest.ApprovalStatus.APPROVED,
                "payment_status": PurchaseRequest.PaymentStatus.PAID,
                "delivery_status": PurchaseRequest.DeliveryStatus.DELIVERED,
                "date_from": "2026-06-01",
                "date_to": "2026-06-15",
            }
        )

        titles = [row[1].value for row in sheet.iter_rows(min_row=2)]
        self.assertEqual(titles, [matching.title])

    def test_purchase_request_xlsx_export_requires_purchase_request_access(self):
        self.login(self.no_access_user)

        response = self.client.get(reverse("purchase_request_export_xlsx"))

        self.assertEqual(response.status_code, 403)

    def test_each_tracking_status_filter_works(self):
        matching = self.create_request(
            title="Filter target",
            order_type=PurchaseRequest.OrderType.EMERGENCY,
            approval_status=PurchaseRequest.ApprovalStatus.APPROVED,
            payment_status=PurchaseRequest.PaymentStatus.SENT_FOR_PAYMENT,
            delivery_status=PurchaseRequest.DeliveryStatus.IN_TRANSIT,
        )
        other = self.create_request(user=self.other_requester, title="Filter other")
        self.login(self.admin)

        filters = {
            "order_type": matching.order_type,
            "approval_status": matching.approval_status,
            "payment_status": matching.payment_status,
            "delivery_status": matching.delivery_status,
        }
        for field, value in filters.items():
            with self.subTest(field=field):
                response = self.client.get(
                    reverse("purchase_request_list"), {field: value}
                )
                self.assertContains(response, matching.title)
                self.assertNotContains(response, other.title)

    def test_search_works_by_item_name_description_and_product_url(self):
        purchase_request = self.create_request(
            title="Hydraulic hose",
            need_description="Repair press line",
            product_url="https://example.com/hose-42",
        )
        self.login(self.admin)

        for query in ["Hydraulic", "press line", "hose-42"]:
            with self.subTest(query=query):
                response = self.client.get(reverse("purchase_request_list"), {"q": query})
                self.assertContains(response, purchase_request.title)

    def test_detail_page_is_visible_to_owner_and_admin_only(self):
        purchase_request = self.create_request()
        url = reverse("purchase_request_detail", args=[purchase_request.pk])

        self.login(self.requester)
        self.assertContains(self.client.get(url), purchase_request.title)
        self.login(self.admin)
        self.assertContains(self.client.get(url), purchase_request.title)
        self.login(self.other_requester)
        self.assertEqual(self.client.get(url).status_code, 404)

    def test_detail_page_shows_audit_fields(self):
        purchase_request = self.create_request(
            status=PurchaseRequest.Status.REJECTED,
            approved_by=self.admin,
            approved_at=timezone.now(),
            rejected_by=self.admin,
            rejected_at=timezone.now(),
            rejection_comment="Budget refused",
        )
        self.login(self.admin)

        response = self.client.get(
            reverse("purchase_request_detail", args=[purchase_request.pk])
        )

        self.assertContains(response, self.requester.username)
        self.assertContains(response, self.admin.username)
        self.assertContains(response, "Budget refused")

    def test_archived_detail_shows_badge_archive_info_and_restore_button(self):
        purchase_request = self.create_request()
        archive_purchase_request(
            purchase_request, archived_by=self.admin, reason="Повністю отримано на склад"
        )
        self.login(self.admin)

        response = self.client.get(
            reverse("purchase_request_detail", args=[purchase_request.pk])
        )

        self.assertContains(response, "В архіві")
        self.assertContains(response, "Повністю отримано на склад")
        self.assertContains(response, "Повернути з архіву")
        self.assertNotContains(response, "Прийняти на склад")

    def test_active_detail_shows_archive_button_for_allowed_user_only(self):
        purchase_request = self.create_request()

        self.login(self.admin)
        response = self.client.get(
            reverse("purchase_request_detail", args=[purchase_request.pk])
        )
        self.assertContains(response, "Архівувати")
        self.assertNotContains(response, "В архіві")

        self.login(self.requester)
        response = self.client.get(
            reverse("purchase_request_detail", args=[purchase_request.pk])
        )
        self.assertNotContains(response, "Архівувати")

    def test_archive_restore_permissions_by_post(self):
        purchase_request = self.create_request()

        self.login(self.requester)
        response = self.client.post(
            reverse("purchase_request_archive_action", args=[purchase_request.pk])
        )
        self.assertEqual(response.status_code, 403)
        purchase_request.refresh_from_db()
        self.assertFalse(purchase_request.is_archived)

        self.login(self.admin)
        response = self.client.post(
            reverse("purchase_request_archive_action", args=[purchase_request.pk]),
            {"archive_reason": "No longer active"},
        )
        self.assertRedirects(
            response, reverse("purchase_request_detail", args=[purchase_request.pk])
        )
        purchase_request.refresh_from_db()
        self.assertTrue(purchase_request.is_archived)
        self.assertEqual(purchase_request.archive_reason, "No longer active")

        self.login(self.requester)
        response = self.client.post(
            reverse("purchase_request_restore_action", args=[purchase_request.pk])
        )
        self.assertEqual(response.status_code, 403)
        purchase_request.refresh_from_db()
        self.assertTrue(purchase_request.is_archived)

        self.login(self.admin)
        response = self.client.post(
            reverse("purchase_request_restore_action", args=[purchase_request.pk])
        )
        self.assertRedirects(
            response, reverse("purchase_request_detail", args=[purchase_request.pk])
        )
        purchase_request.refresh_from_db()
        self.assertFalse(purchase_request.is_archived)

    def test_detail_page_uses_compact_two_column_layout(self):
        purchase_request = self.create_request(
            unit_price_uah=Decimal("55.50"),
            requested_qty=Decimal("12.000"),
        )
        self.login(self.admin)

        response = self.client.get(
            reverse("purchase_request_detail", args=[purchase_request.pk])
        )

        self.assertContains(response, "purchase-detail-grid")
        self.assertContains(response, "purchase-detail-main")
        self.assertContains(response, "purchase-detail-side")
        self.assertContains(response, "purchase-detail-status-panel")
        self.assertContains(response, "purchase-detail-fields--status-top")
        self.assertContains(response, "purchase-detail-status-form")
        self.assertContains(response, "purchase-detail-actions-panel")
        self.assertContains(response, "purchase-detail-actions--top")
        self.assertContains(response, "purchase-detail-action-details")
        self.assertContains(response, "purchase-detail-actions--primary")
        self.assertContains(response, "purchase-detail-metrics")
        self.assertContains(response, "detail-payment-status")
        self.assertContains(response, "detail-delivery-status")
        self.assertContains(response, "Кількість / одиниця")
        self.assertContains(response, "12,000 pairs")
        self.assertContains(response, "Вартість за одиницю")
        self.assertContains(response, "Сума")
        html = response.content.decode()
        self.assertLess(
            html.index("purchase-detail-actions-panel"),
            html.index("purchase-detail-grid"),
        )
        status_panel_start = html.index("purchase-detail-status-panel")
        self.assertNotIn(
            "purchase-detail-actions--primary",
            html[status_panel_start:],
        )

    def test_owner_can_edit_draft_but_not_submitted_request(self):
        draft = self.create_request()
        self.login(self.requester)

        response = self.client.post(
            reverse("purchase_request_update", args=[draft.pk]),
            self.request_data(title="Updated title"),
        )
        draft.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(draft.title, "Updated title")

        draft.status = PurchaseRequest.Status.PENDING_APPROVAL
        draft.save(update_fields=["status"])
        self.assertEqual(
            self.client.get(reverse("purchase_request_update", args=[draft.pk])).status_code,
            404,
        )

    def test_owner_can_send_draft_for_approval(self):
        purchase_request = self.create_request()

        response = self.post_action(purchase_request, "send", self.requester)

        purchase_request.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(purchase_request.status, PurchaseRequest.Status.PENDING_APPROVAL)

    def test_admin_can_approve_pending_request(self):
        purchase_request = self.create_request(status=PurchaseRequest.Status.PENDING_APPROVAL)

        self.post_action(purchase_request, "approve")

        purchase_request.refresh_from_db()
        self.assertEqual(purchase_request.status, PurchaseRequest.Status.APPROVED)
        self.assertEqual(
            purchase_request.approval_status,
            PurchaseRequest.ApprovalStatus.APPROVED,
        )
        self.assertEqual(purchase_request.approved_by, self.admin)
        self.assertIsNotNone(purchase_request.approved_at)

    def test_admin_can_reject_pending_request(self):
        purchase_request = self.create_request(status=PurchaseRequest.Status.PENDING_APPROVAL)

        self.login(self.admin)
        self.client.post(
            reverse("purchase_request_reject", args=[purchase_request.pk]),
            {"rejection_comment": "Too expensive"},
        )

        purchase_request.refresh_from_db()
        self.assertEqual(purchase_request.status, PurchaseRequest.Status.REJECTED)
        self.assertEqual(
            purchase_request.approval_status,
            PurchaseRequest.ApprovalStatus.REJECTED,
        )
        self.assertEqual(purchase_request.rejected_by, self.admin)
        self.assertIsNotNone(purchase_request.rejected_at)
        self.assertEqual(purchase_request.rejection_comment, "Too expensive")

    def test_admin_can_mark_approved_request_as_ordered(self):
        purchase_request = self.create_request(status=PurchaseRequest.Status.APPROVED)

        self.post_action(purchase_request, "order")

        purchase_request.refresh_from_db()
        self.assertEqual(purchase_request.status, PurchaseRequest.Status.ORDERED)

    def test_admin_can_cancel_request(self):
        purchase_request = self.create_request(status=PurchaseRequest.Status.ORDERED)

        self.post_action(purchase_request, "cancel")

        purchase_request.refresh_from_db()
        self.assertEqual(purchase_request.status, PurchaseRequest.Status.CANCELLED)

    def test_regular_user_cannot_approve_reject_order_or_cancel(self):
        transitions = [
            ("approve", PurchaseRequest.Status.PENDING_APPROVAL),
            ("reject", PurchaseRequest.Status.PENDING_APPROVAL),
            ("order", PurchaseRequest.Status.APPROVED),
            ("cancel", PurchaseRequest.Status.DRAFT),
        ]
        for action, status in transitions:
            with self.subTest(action=action):
                purchase_request = self.create_request(status=status)
                response = self.post_action(purchase_request, action, self.requester)
                purchase_request.refresh_from_db()
                self.assertEqual(response.status_code, 403)
                self.assertEqual(purchase_request.status, status)

    def test_user_without_warehouse_access_cannot_list_or_create(self):
        self.login(self.no_access_user)

        self.assertEqual(self.client.get(reverse("purchase_request_list")).status_code, 403)
        self.assertEqual(self.client.get(reverse("purchase_request_create")).status_code, 403)

    def test_user_with_view_permission_can_see_purchase_request_list(self):
        purchase_request = self.create_request()
        self.grant_permission(self.no_access_user, "can_view_purchase_requests")
        self.login(self.no_access_user)

        response = self.client.get(reverse("purchase_request_list"))

        self.assertContains(response, purchase_request.title)

    def test_user_with_create_permission_can_create_purchase_request(self):
        self.grant_permission(self.no_access_user, "can_create_purchase_requests")
        self.login(self.no_access_user)

        response = self.client.post(
            reverse("purchase_request_create"), self.request_data()
        )

        purchase_request = PurchaseRequest.objects.get(requested_by=self.no_access_user)
        self.assertRedirects(
            response, reverse("purchase_request_detail", args=[purchase_request.pk])
        )

    def test_user_without_create_permission_cannot_create_by_post(self):
        self.login(self.no_access_user)

        response = self.client.post(
            reverse("purchase_request_create"), self.request_data()
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(PurchaseRequest.objects.filter(requested_by=self.no_access_user).exists())

    def test_approver_permission_controls_buttons_and_direct_post(self):
        purchase_request = self.create_request(
            status=PurchaseRequest.Status.PENDING_APPROVAL
        )
        self.grant_permission(self.other_requester, "can_view_purchase_requests")
        self.login(self.other_requester)

        response = self.client.get(
            reverse("purchase_request_detail", args=[purchase_request.pk])
        )
        self.assertNotContains(response, "Погодити")

        denied = self.client.post(
            reverse("purchase_request_approve", args=[purchase_request.pk])
        )
        self.assertEqual(denied.status_code, 403)

        self.grant_permission(self.other_requester, "can_approve_purchase_requests")
        response = self.client.get(
            reverse("purchase_request_detail", args=[purchase_request.pk])
        )
        self.assertContains(response, "Погодити")

        approved = self.client.post(
            reverse("purchase_request_approve", args=[purchase_request.pk])
        )
        purchase_request.refresh_from_db()
        self.assertEqual(approved.status_code, 302)
        self.assertEqual(purchase_request.approved_by, self.other_requester)
        self.assertIsNotNone(purchase_request.approved_at)

    def test_approver_permission_can_approve_own_request(self):
        purchase_request = self.create_request(
            user=self.requester,
            status=PurchaseRequest.Status.PENDING_APPROVAL,
        )
        self.grant_permission(self.requester, "can_approve_purchase_requests")

        response = self.post_action(purchase_request, "approve", self.requester)

        purchase_request.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(purchase_request.approved_by, self.requester)

    def test_approver_permission_can_reject_and_writes_audit(self):
        purchase_request = self.create_request(
            status=PurchaseRequest.Status.PENDING_APPROVAL
        )
        self.grant_permission(self.other_requester, "can_view_purchase_requests")
        self.grant_permission(self.other_requester, "can_approve_purchase_requests")
        self.login(self.other_requester)

        response = self.client.post(
            reverse("purchase_request_reject", args=[purchase_request.pk]),
            {"rejection_comment": "Not enough budget"},
        )

        purchase_request.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(purchase_request.rejected_by, self.other_requester)
        self.assertIsNotNone(purchase_request.rejected_at)
        self.assertEqual(purchase_request.rejection_comment, "Not enough budget")

    def test_tracking_permission_controls_payment_and_delivery_update(self):
        purchase_request = self.create_request()
        self.grant_permission(self.other_requester, "can_view_purchase_requests")
        self.login(self.other_requester)

        denied = self.client.post(
            reverse("purchase_request_tracking_status", args=[purchase_request.pk]),
            {"payment_status": PurchaseRequest.PaymentStatus.PAID},
        )
        self.assertEqual(denied.status_code, 403)

        self.grant_permission(
            self.other_requester, "can_update_purchase_request_tracking"
        )
        movement_count = StockMovement.objects.count()
        balance_count = StockBalance.objects.count()
        updated = self.client.post(
            reverse("purchase_request_tracking_status", args=[purchase_request.pk]),
            {
                "payment_status": PurchaseRequest.PaymentStatus.PAID,
                "delivery_status": PurchaseRequest.DeliveryStatus.DELIVERED,
            },
        )

        purchase_request.refresh_from_db()
        self.assertEqual(updated.status_code, 302)
        self.assertEqual(purchase_request.payment_status, PurchaseRequest.PaymentStatus.PAID)
        self.assertEqual(
            purchase_request.delivery_status, PurchaseRequest.DeliveryStatus.DELIVERED
        )
        self.assertEqual(StockMovement.objects.count(), movement_count)
        self.assertEqual(StockBalance.objects.count(), balance_count)

    def test_purchase_request_export_works_for_view_permission_user(self):
        purchase_request = self.create_request(title="Permission export")
        self.grant_permission(self.no_access_user, "can_view_purchase_requests")

        response, sheet = self.load_purchase_request_export(user=self.no_access_user)

        self.assertEqual(response.status_code, 200)
        titles = [row[1].value for row in sheet.iter_rows(min_row=2)]
        self.assertIn(purchase_request.title, titles)

    def test_invalid_status_transition_is_rejected(self):
        purchase_request = self.create_request(status=PurchaseRequest.Status.DRAFT)

        response = self.post_action(purchase_request, "approve")

        purchase_request.refresh_from_db()
        self.assertEqual(response.status_code, 400)
        self.assertEqual(purchase_request.status, PurchaseRequest.Status.DRAFT)

    def test_purchase_workflow_does_not_create_movements_or_balances(self):
        purchase_request = self.create_request()
        movement_count = StockMovement.objects.count()
        balance_count = StockBalance.objects.count()

        self.post_action(purchase_request, "send", self.requester)
        self.post_action(purchase_request, "approve")
        self.post_action(purchase_request, "order")

        self.assertEqual(StockMovement.objects.count(), movement_count)
        self.assertEqual(StockBalance.objects.count(), balance_count)

    def test_regular_user_cannot_manage_tracking_statuses(self):
        purchase_request = self.create_request()
        self.login(self.requester)
        data = self.request_data(
            approval_status=PurchaseRequest.ApprovalStatus.APPROVED,
            payment_status=PurchaseRequest.PaymentStatus.PAID,
            delivery_status=PurchaseRequest.DeliveryStatus.DELIVERED,
        )

        self.client.post(
            reverse("purchase_request_update", args=[purchase_request.pk]), data
        )

        purchase_request.refresh_from_db()
        self.assertEqual(
            purchase_request.approval_status, PurchaseRequest.ApprovalStatus.PENDING
        )
        self.assertEqual(
            purchase_request.payment_status,
            PurchaseRequest.PaymentStatus.INVOICE_NOT_RECEIVED,
        )
        self.assertEqual(
            purchase_request.delivery_status,
            PurchaseRequest.DeliveryStatus.NOT_SHIPPED,
        )

    def test_admin_can_update_payment_and_delivery_without_stock_effects(self):
        purchase_request = self.create_request()
        self.login(self.admin)
        movement_count = StockMovement.objects.count()
        balance_count = StockBalance.objects.count()

        response = self.client.post(
            reverse("purchase_request_update", args=[purchase_request.pk]),
            self.request_data(
                approval_status=PurchaseRequest.ApprovalStatus.PENDING,
                payment_status=PurchaseRequest.PaymentStatus.PAID,
                delivery_status=PurchaseRequest.DeliveryStatus.DELIVERED,
            ),
        )

        purchase_request.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            purchase_request.approval_status, PurchaseRequest.ApprovalStatus.PENDING
        )
        self.assertEqual(purchase_request.payment_status, PurchaseRequest.PaymentStatus.PAID)
        self.assertEqual(
            purchase_request.delivery_status, PurchaseRequest.DeliveryStatus.DELIVERED
        )
        self.assertEqual(StockMovement.objects.count(), movement_count)
        self.assertEqual(StockBalance.objects.count(), balance_count)

    def test_edit_page_cannot_update_approval_status_or_audit_fields(self):
        purchase_request = self.create_request(
            approved_by=self.requester,
            approved_at=timezone.now(),
            rejected_by=self.other_requester,
            rejected_at=timezone.now(),
            rejection_comment="Original rejection",
        )
        original_requested_by = purchase_request.requested_by
        original_approved_by = purchase_request.approved_by
        original_approved_at = purchase_request.approved_at
        original_rejected_by = purchase_request.rejected_by
        original_rejected_at = purchase_request.rejected_at
        self.login(self.admin)

        response = self.client.post(
            reverse("purchase_request_update", args=[purchase_request.pk]),
            self.request_data(
                requested_by=self.other_requester.pk,
                approval_status=PurchaseRequest.ApprovalStatus.APPROVED,
                payment_status=PurchaseRequest.PaymentStatus.INVOICE_NOT_RECEIVED,
                delivery_status=PurchaseRequest.DeliveryStatus.NOT_SHIPPED,
                approved_by=self.admin.pk,
                approved_at=timezone.now().isoformat(),
                rejected_by=self.admin.pk,
                rejected_at=timezone.now().isoformat(),
                rejection_comment="Tampered",
            ),
        )

        purchase_request.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            purchase_request.approval_status, PurchaseRequest.ApprovalStatus.PENDING
        )
        self.assertEqual(purchase_request.status, PurchaseRequest.Status.DRAFT)
        self.assertEqual(purchase_request.requested_by, original_requested_by)
        self.assertEqual(purchase_request.approved_by, original_approved_by)
        self.assertEqual(purchase_request.approved_at, original_approved_at)
        self.assertEqual(purchase_request.rejected_by, original_rejected_by)
        self.assertEqual(purchase_request.rejected_at, original_rejected_at)
        self.assertEqual(purchase_request.rejection_comment, "Original rejection")

    def test_sheet_fields_do_not_include_code_or_invoice_fields(self):
        self.login(self.requester)

        response = self.client.get(reverse("purchase_request_create"))

        fields = response.context["form"].fields
        self.assertNotIn("code", fields)
        self.assertNotIn("item_code", fields)
        self.assertNotIn("order_number", fields)
        self.assertNotIn("invoice_number", fields)

    def test_purchase_list_title_is_localized(self):
        expected_titles = {
            "uk": "Заявки на закупівлю",
            "ru": "Заявки на закупку",
            "en": "Purchase requests",
            "pl": "Wnioski zakupowe",
            "it": "Richieste di acquisto",
        }
        self.login(self.admin)

        for language, title in expected_titles.items():
            with self.subTest(language=language), translation.override(language):
                response = self.client.get(f"/{language}/purchases/")
                self.assertContains(response, title)

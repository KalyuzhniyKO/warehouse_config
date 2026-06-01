import csv
import json
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.forms.models import model_to_dict
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.html import format_html, linebreaks, urlize
from django.utils.safestring import mark_safe
from django.utils.translation import gettext_lazy as _
from itertools import groupby
from django.views.generic import CreateView, FormView, ListView, TemplateView, UpdateView, View

from ..forms import (
    AnalyticsFilterForm,
    CategoryForm,
    ItemForm,
    InitialBalanceForm,
    InventoryCountCreateForm,
    InventoryCountLineForm,
    LabelTemplateForm,
    LocationForm,
    PrintLabelForm,
    PrinterForm,
    RecipientForm,
    StockBalanceFilterForm,
    StockIssueForm,
    StockMovementFilterForm,
    StockReceiveForm,
    StockTransferForm,
    UnitForm,
    WarehouseForm,
)
from ..models import (
    Category,
    InventoryCount,
    Item,
    LabelTemplate,
    Location,
    PrintJob,
    Printer,
    Recipient,
    StockBalance,
    StockMovement,
    Unit,
    Warehouse,
)
from ..permissions import (
    ANALYTICS_GROUPS,
    MANAGEMENT_GROUPS,
    DIRECTORY_EDIT_GROUPS,
    PRINT_GROUPS,
    SETTINGS_GROUPS,
    STOCK_EDIT_GROUPS,
    STOCK_VIEW_GROUPS,
    USER_MANAGEMENT_GROUPS,
    GroupRequiredMixin,
    user_in_groups,
)
from ..services import analytics as analytics_service
from ..services.analytics_presets import get_analytics_report_presets
from ..services.filter_memory import apply_remembered_filters, build_redirect_url, querydict_from_params
from ..services.inventory import (
    InventoryServiceError,
    complete_inventory_count,
    create_inventory_count,
    update_inventory_line_actual_qty,
)
from ..services.labels import download_item_label_pdf, get_default_label_template, print_item_label
from ..services.warehouse_access import get_accessible_warehouses
from ..services.stock import (
    InsufficientStockError,
    SameLocationTransferError,
    StockServiceError,
    create_initial_balance,
    issue_stock,
    receive_stock,
    transfer_stock,
)






DATA_QUALITY_CHECK_META = [
    {"key": "missing_documents", "label": _("Рухи без документа"), "description": _("Рухи, у яких відсутній номер документа."), "journal_param": "no_document=1"},
    {"key": "issue_without_recipient", "label": _("Видача без отримувача"), "description": _("Операції видачі без вказаного отримувача."), "journal_param": "missing_recipient=1"},
    {"key": "issue_without_usage_place", "label": _("Видача без цеху / місця використання"), "description": _("Операції видачі без заповненого цеху або місця використання."), "journal_param": "missing_usage_place=1"},
    {"key": "movement_without_item", "label": _("Рухи без товару"), "description": _("Операції, у яких не вказано товар."), "journal_param": "missing_item=1"},
    {"key": "non_positive_qty", "label": _("Некоректна кількість"), "description": _("Операції з нульовою або від’ємною кількістю."), "journal_param": "non_positive_qty=1"},
    {"key": "receive_without_destination", "label": _("Прихід без складу призначення"), "description": _("Операції приходу без складу або локації призначення."), "journal_param": "missing_destination=1"},
]

class ManagementReportsView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    group_names = ANALYTICS_GROUPS
    template_name = "core/management/reports.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        presets = get_analytics_report_presets()
        grouped = []
        for _, items in groupby(presets, key=lambda x: x["category"]):
            items = list(items)
            grouped.append({"title": items[0]["category_title"], "presets": items})
        context["preset_groups"] = grouped
        return context
class AnalyticsRedirectView(LoginRequiredMixin, GroupRequiredMixin, View):
    group_names = ANALYTICS_GROUPS

    def get(self, request, *args, **kwargs):
        return redirect("management_analytics")


def clean_analytics_filters(form, user=None):
    filters = {}
    if user is not None:
        filters["accessible_warehouses"] = get_accessible_warehouses(user)
    if not form.is_valid():
        return filters
    filters.update(
        {key: value for key, value in form.cleaned_data.items() if value not in (None, "")}
    )
    period = form.cleaned_data.get("period")
    if period:
        filters.update(analytics_service.get_analytics_filters(form.cleaned_data))
    return filters


def build_top_issued_items_visual(top_issued_items):
    max_qty = max((row.get("total_qty") or 0 for row in top_issued_items), default=0)
    visual_rows = []
    for row in top_issued_items:
        total_qty = row.get("total_qty") or 0
        bar_percent = int((total_qty / max_qty) * 100) if max_qty else 0
        visual_rows.append(
            {
                "item_id": row.get("item__id"),
                "item_name": row.get("item__name"),
                "total_qty": total_qty,
                "bar_percent": bar_percent,
            }
        )
    return visual_rows


def build_operation_mix_visual(operation_mix):
    labels = {
        "receive": _("Прихід"),
        "issue": _("Видача"),
        "return": _("Повернення"),
        "write_off": _("Списання"),
        "transfer": _("Переміщення"),
    }
    rows_by_key = {row.get("key"): row for row in operation_mix}
    return [
        {
            "key": key,
            "label": label,
            "total": rows_by_key.get(key, {}).get("total", 0) or 0,
            "bar_percent": int(rows_by_key.get(key, {}).get("percent", 0) or 0),
        }
        for key, label in labels.items()
    ]


class AnalyticsView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    group_names = ANALYTICS_GROUPS
    template_name = "core/management/analytics.html"
    page_key = "management_analytics"

    def get(self, request, *args, **kwargs):
        params, used_remembered_filters, should_redirect = apply_remembered_filters(request, self.page_key)
        self.used_remembered_filters = used_remembered_filters
        if should_redirect:
            return redirect(build_redirect_url(request.path, params))
        self.effective_get = querydict_from_params(params) if params else request.GET
        return super().get(request, *args, **kwargs)

    def get_filter_form(self):
        return AnalyticsFilterForm(
            getattr(self, "effective_get", self.request.GET) or None,
            request_user=self.request.user,
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = self.get_filter_form()
        filters = clean_analytics_filters(form, self.request.user)
        summary = analytics_service.get_analytics_summary(filters)
        previous_filters = analytics_service.get_previous_period(filters)
        previous_summary = analytics_service.get_analytics_summary(previous_filters) if previous_filters.get("date_from") else None
        summary_deltas = {}
        for key in ["operations_count", "receive_qty", "issue_qty", "return_qty", "writeoff_qty"]:
            summary_deltas[key] = analytics_service.get_kpi_delta(summary.get(key), previous_summary.get(key) if previous_summary else None)

        request_get = getattr(self, "effective_get", self.request.GET)
        query_base = {k: v for k, v in request_get.items() if v}
        date_params = {"date_from": filters.get("date_from"), "date_to": filters.get("date_to")}
        movement_query = {**{k: str(v) for k, v in query_base.items() if k in {"warehouse", "location", "movement_type"}}, **{k: str(v) for k, v in date_params.items() if v}}
        filter_query = analytics_service.build_analytics_filter_query(filters)
        daily_movement = analytics_service.get_daily_movement(filters)
        top_issued_items = analytics_service.get_top_issued_items(filters)
        operation_mix = analytics_service.get_operation_mix(filters)
        top_usage_places = analytics_service.get_top_usage_places(filters)
        top_recipients = analytics_service.get_top_recipients(filters)
        data_quality = analytics_service.get_analytics_data_quality(filters)
        context.update(
            {
                "filter_form": form,
                "summary": summary,
                "summary_deltas": summary_deltas,
                "daily_movement": daily_movement,
                "operation_mix": operation_mix,
                "operation_mix_visual": build_operation_mix_visual(operation_mix),
                "top_issued_items": top_issued_items,
                "top_issued_items_visual": build_top_issued_items_visual(top_issued_items),
                "top_usage_places": top_usage_places,
                "top_recipients": top_recipients,
                "inactive_stock_items": analytics_service.get_inactive_stock_items(filters),
                "recent_movements": analytics_service.get_recent_movements(filters),
                "movement_list_base_url": reverse("movement_list"),
                "movement_query": urlencode(movement_query),
                "filter_query": urlencode(filter_query),
                "quick_periods": [("today", _("Сьогодні")), ("7d", _("7 днів")), ("30d", _("30 днів")), ("month", _("Поточний місяць")), ("prev_month", _("Попередній місяць"))],
                "data_quality": data_quality,
                "data_quality_url": reverse("management_analytics_data_quality") + "?" + urlencode(filter_query),
                "used_remembered_filters": getattr(self, "used_remembered_filters", False),
            }
        )
        return context


movement_export_location = analytics_service.movement_export_location


class AnalyticsCSVExportView(LoginRequiredMixin, GroupRequiredMixin, View):
    group_names = ANALYTICS_GROUPS

    def get(self, request, *args, **kwargs):
        form = AnalyticsFilterForm(request.GET or None, request_user=request.user)
        filters = clean_analytics_filters(form, request.user)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            'attachment; filename="warehouse-analytics.csv"'
        )
        response.write("\ufeff")
        writer = csv.writer(response)
        for section_index, (header, rows) in enumerate(analytics_service.get_csv_export_sections(filters)):
            if section_index:
                writer.writerow([])
            writer.writerow(header)
            for row in rows:
                writer.writerow(row)
        return response


class AnalyticsXLSXExportView(LoginRequiredMixin, GroupRequiredMixin, View):
    group_names = ANALYTICS_GROUPS

    def get(self, request, *args, **kwargs):
        try:
            from openpyxl import Workbook
        except ImportError:
            return HttpResponse(_("XLSX експорт недоступний: openpyxl не встановлено."), status=501)
        form = AnalyticsFilterForm(request.GET or None, request_user=request.user)
        filters = clean_analytics_filters(form, request.user)
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Metric", "Value"])
        for row in analytics_service.get_xlsx_summary_rows(filters):
            ws.append(row)
        for title, headers, rows in analytics_service.get_xlsx_data_sheets(filters):
            sh = wb.create_sheet(title)
            sh.append(headers)
            for row in rows:
                sh.append(row)
            sh.freeze_panes = 'A2'
            sh.auto_filter.ref = sh.dimensions
        dq_sheet = wb.create_sheet("Data quality")
        dq_sheet.append(["Metric", "Value"])
        metric_rows, check_rows = analytics_service.get_xlsx_data_quality_rows(filters)
        for row in metric_rows:
            dq_sheet.append(row)
        dq_sheet.append([])
        dq_sheet.append(["Check", "Count"])
        for row in check_rows:
            dq_sheet.append(row)
        dq_sheet.freeze_panes = 'A2'
        dq_sheet.auto_filter.ref = dq_sheet.dimensions
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="warehouse-analytics.xlsx"'
        wb.save(response)
        return response




class AnalyticsDataQualityView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    group_names = ANALYTICS_GROUPS
    template_name = "core/management/analytics_data_quality.html"
    page_key = "management_analytics_data_quality"

    def get(self, request, *args, **kwargs):
        params, used_remembered_filters, should_redirect = apply_remembered_filters(request, self.page_key)
        self.used_remembered_filters = used_remembered_filters
        if should_redirect:
            return redirect(build_redirect_url(request.path, params))
        self.effective_get = querydict_from_params(params) if params else request.GET
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = AnalyticsFilterForm(
            getattr(self, "effective_get", self.request.GET) or None,
            request_user=self.request.user,
        )
        filters = clean_analytics_filters(form, self.request.user)
        filter_query = urlencode(analytics_service.build_analytics_filter_query(filters))
        data_quality = analytics_service.get_analytics_data_quality(filters)
        quality_checks = []
        for meta in DATA_QUALITY_CHECK_META:
            check = data_quality["checks"].get(meta["key"], {"count": 0, "examples": []})
            quality_checks.append({
                "key": meta["key"], "label": meta["label"], "description": meta["description"],
                "count": check["count"], "examples": check["examples"],
                "journal_url": f"{reverse('movement_list')}?{filter_query}&{meta['journal_param']}",
            })
        context.update({
            "filter_form": form,
            "data_quality": data_quality,
            "quality_checks": quality_checks,
            "filter_query": filter_query,
            "movement_list_base_url": reverse("movement_list"),
            "stock_checked_note": _("Залишки перевіряються на поточний момент."),
            "used_remembered_filters": getattr(self, "used_remembered_filters", False),
        })
        return context
class AnalyticsItemDetailView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    group_names = ANALYTICS_GROUPS
    template_name = "core/management/analytics_item_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = AnalyticsFilterForm(self.request.GET or None, request_user=self.request.user)
        filters = clean_analytics_filters(form, self.request.user)
        item = get_object_or_404(Item, pk=self.kwargs['item_id'])
        data = analytics_service.get_item_analytics(item, filters)
        filter_query = urlencode(analytics_service.build_analytics_filter_query(filters))
        context.update({'item': item, 'filter_form': form, 'filters': filters, 'data': data, 'movement_list_base_url': reverse('movement_list'), 'filter_query': filter_query})
        return context


class AnalyticsUsagePlaceDetailView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    group_names = ANALYTICS_GROUPS
    template_name = "core/management/analytics_usage_place_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = AnalyticsFilterForm(self.request.GET or None, request_user=self.request.user)
        filters = clean_analytics_filters(form, self.request.user)
        usage_place = self.kwargs['usage_place_id']
        data = analytics_service.get_usage_place_analytics(usage_place, filters)
        filter_query = urlencode(analytics_service.build_analytics_filter_query(filters))
        context.update({'usage_place': usage_place, 'filters': filters, 'data': data, 'movement_list_base_url': reverse('movement_list'), 'filter_query': filter_query})
        return context


class AnalyticsRecipientDetailView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    group_names = ANALYTICS_GROUPS
    template_name = "core/management/analytics_recipient_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = AnalyticsFilterForm(self.request.GET or None, request_user=self.request.user)
        filters = clean_analytics_filters(form, self.request.user)
        recipient = get_object_or_404(Recipient, pk=self.kwargs['recipient_id'])
        data = analytics_service.get_recipient_analytics(recipient, filters)
        filter_query = urlencode(analytics_service.build_analytics_filter_query(filters))
        context.update({'recipient': recipient, 'filters': filters, 'data': data, 'movement_list_base_url': reverse('movement_list'), 'filter_query': filter_query})
        return context

import csv
import json
from datetime import datetime
from html import escape
from io import BytesIO
from pathlib import Path

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.forms.models import model_to_dict
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.formats import date_format
from django.utils.html import format_html, linebreaks, urlize
from django.utils.safestring import mark_safe
from django.utils.translation import gettext, gettext_lazy as _
from django.views.generic import CreateView, FormView, ListView, TemplateView, UpdateView, View

from ..forms import (
    AnalyticsFilterForm,
    CategoryForm,
    ItemForm,
    InitialBalanceForm,
    InventoryCountCreateForm,
    InventoryCountLineForm,
    LabelTemplateForm,
    LocationForm,
    PrintLabelForm,
    PrinterForm,
    RecipientForm,
    StockBalanceFilterForm,
    StockIssueForm,
    StockMovementFilterForm,
    StockReceiveForm,
    StockTransferForm,
    UnitForm,
    WarehouseForm,
)
from ..models import (
    Category,
    InventoryCount,
    Item,
    LabelTemplate,
    Location,
    PrintJob,
    Printer,
    Recipient,
    StockBalance,
    StockMovement,
    Unit,
    Warehouse,
)
from ..permissions import (
    ANALYTICS_GROUPS,
    MANAGEMENT_GROUPS,
    DIRECTORY_EDIT_GROUPS,
    PRINT_GROUPS,
    SETTINGS_GROUPS,
    STOCK_EDIT_GROUPS,
    STOCK_VIEW_GROUPS,
    USER_MANAGEMENT_GROUPS,
    GroupRequiredMixin,
    user_in_groups,
)
from ..services import analytics as analytics_service
from ..services.inventory import (
    InventoryServiceError,
    complete_inventory_count,
    create_inventory_count,
    reconcile_inventory_line,
    update_inventory_line_actual_qty,
)
from ..services.labels import download_item_label_pdf, get_default_label_template, print_item_label
from ..services.barcodes import resolve_item_barcode
from ..services.warehouse_access import get_accessible_warehouses
from ..services.stock import (
    InsufficientStockError,
    SameLocationTransferError,
    StockServiceError,
    create_initial_balance,
    issue_stock,
    receive_stock,
    transfer_stock,
)



def get_inventory_export_queryset(inventory_count):
    return inventory_count.lines.select_related(
        "inventory_count__warehouse",
        "item",
        "item__barcode",
        "item__unit",
        "location",
        "location__warehouse",
        "counted_by",
    ).order_by("item__name", "location__name", "id")


def inventory_line_export_row(inventory_count, line):
    reconcile_inventory_line(line)
    return [
        inventory_count.number,
        inventory_count.get_status_display(),
        inventory_count.warehouse.name,
        inventory_count.location.name if inventory_count.location else "",
        line.item.name,
        line.item.internal_code,
        line.barcode or (line.item.barcode.barcode if line.item.barcode_id else ""),
        line.location.name,
        line.snapshot_qty,
        line.movement_delta,
        line.expected_qty_at_count_time,
        line.counted_at or "",
        line.actual_qty if line.actual_qty is not None else "",
        line.variance_qty,
        line.comment,
    ]


INVENTORY_EXPORT_HEADERS = [
    _("Номер інвентаризації"),
    _("Статус"),
    _("Склад"),
    _("Локація інвентаризації"),
    _("Номенклатура"),
    _("Internal code"),
    _("Barcode"),
    _("Локація"),
    _("Знімок на початок"),
    _("Рухи під час інвентаризації"),
    _("Очікувана кількість на час підрахунку"),
    _("Час підрахунку"),
    _("Фактична кількість"),
    _("Відхилення"),
    _("Коментар рядка"),
]

INVENTORY_RESULT_HEADERS = [
    _("Код товару"),
    _("Штрихкод"),
    _("Назва товару"),
    _("Одиниця"),
    _("Знімок на початок"),
    _("Рухи під час інвентаризації"),
    _("Очікувана кількість на час підрахунку"),
    _("Фактична кількість"),
    _("Відхилення"),
    _("Порахував"),
    _("Час підрахунку"),
    _("Коментар рядка"),
]


def get_completed_inventory_for_export(user, pk):
    return get_object_or_404(
        InventoryCount.objects.select_related(
            "warehouse", "location", "created_by"
        ).filter(
            warehouse__in=get_accessible_warehouses(user),
            status=InventoryCount.Status.COMPLETED,
        ),
        pk=pk,
    )


def get_inventory_result_lines(inventory_count):
    lines = list(get_inventory_export_queryset(inventory_count))
    for line in lines:
        reconcile_inventory_line(line)
    return lines


def inventory_user_display_name(user):
    if not user:
        return ""
    return user.get_full_name() or user.get_username()


def inventory_export_summary(inventory_count, lines, generated_at):
    counted_lines = [line for line in lines if line.actual_qty is not None]
    variances = [line.variance_qty for line in counted_lines]
    return [
        (_("Номер інвентаризації"), inventory_count.number),
        (_("Склад"), inventory_count.warehouse.name),
        (_("Статус"), inventory_count.get_status_display()),
        (_("Створив"), inventory_user_display_name(inventory_count.created_by)),
        (_("Дата початку"), inventory_count.started_at),
        (_("Дата завершення"), inventory_count.completed_at),
        (_("Всього рядків"), len(lines)),
        (_("Всього підраховано"), len(counted_lines)),
        (_("Кількість нестач"), sum(variance < 0 for variance in variances)),
        (_("Кількість надлишків"), sum(variance > 0 for variance in variances)),
        (
            _("Сума нестач"),
            abs(sum((variance for variance in variances if variance < 0), start=0)),
        ),
        (
            _("Сума надлишків"),
            sum((variance for variance in variances if variance > 0), start=0),
        ),
        (_("Згенеровано"), generated_at),
    ]


def inventory_result_row(line):
    return [
        line.item.internal_code,
        line.barcode or (line.item.barcode.barcode if line.item.barcode_id else ""),
        line.item.name,
        line.item.unit.symbol,
        line.snapshot_qty,
        line.movement_delta,
        line.expected_qty_at_count_time,
        line.actual_qty if line.actual_qty is not None else "",
        line.variance_qty,
        inventory_user_display_name(line.counted_by),
        line.counted_at or "",
        line.comment,
    ]


def local_naive_datetime(value):
    if isinstance(value, datetime) and timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    return value


class InventoryCSVExportView(LoginRequiredMixin, GroupRequiredMixin, View):
    group_names = STOCK_VIEW_GROUPS

    def get(self, request, pk, *args, **kwargs):
        inventory_count = get_object_or_404(
            InventoryCount.objects.select_related("warehouse", "location").filter(
                warehouse__in=get_accessible_warehouses(request.user)
            ),
            pk=pk,
        )
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="inventory_{inventory_count.number}.csv"'
        )
        response.write("\ufeff")
        writer = csv.writer(response)
        writer.writerow(INVENTORY_EXPORT_HEADERS)
        for line in get_inventory_export_queryset(inventory_count):
            writer.writerow(inventory_line_export_row(inventory_count, line))
        return response


class InventoryXLSXExportView(LoginRequiredMixin, GroupRequiredMixin, View):
    group_names = STOCK_VIEW_GROUPS

    def get(self, request, pk, *args, **kwargs):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return HttpResponse(_("XLSX export requires openpyxl."), status=503)

        inventory_count = get_completed_inventory_for_export(request.user, pk)
        lines = get_inventory_result_lines(inventory_count)
        generated_at = timezone.now()
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.append([gettext("Підсумок інвентаризації")])
        for label, value in inventory_export_summary(
            inventory_count, lines, generated_at
        ):
            summary.append([str(label), local_naive_datetime(value)])
        for cell in summary["A"]:
            cell.font = Font(bold=True, color="92400E")
        summary["A1"].font = Font(bold=True, size=14, color="92400E")
        for row_number in (6, 7, 14):
            summary.cell(row=row_number, column=2).number_format = "yyyy-mm-dd hh:mm:ss"
        for row_number in (12, 13):
            summary.cell(row=row_number, column=2).number_format = "#,##0.000"
        summary.column_dimensions["A"].width = 34
        summary.column_dimensions["B"].width = 44

        results = workbook.create_sheet("Results")
        results.append([str(header) for header in INVENTORY_RESULT_HEADERS])
        header_fill = PatternFill("solid", fgColor="D4AC00")
        for cell in results[1]:
            cell.font = Font(bold=True, color="101828")
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for line in lines:
            row = inventory_result_row(line)
            results.append(
                [
                    *row[:4],
                    *[float(value) if value != "" else "" for value in row[4:9]],
                    row[9],
                    local_naive_datetime(row[10]),
                    row[11],
                ]
            )

        widths = [20, 22, 36, 14, 20, 22, 24, 20, 18, 24, 21, 38]
        for index, width in enumerate(widths, start=1):
            column_letter = results.cell(row=1, column=index).column_letter
            results.column_dimensions[column_letter].width = width
        results.freeze_panes = "A2"
        results.auto_filter.ref = results.dimensions
        for row in results.iter_rows(min_row=2, min_col=5, max_col=9):
            for cell in row:
                cell.number_format = "#,##0.000"
        for cell in results["K"][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="inventory_{inventory_count.number}.xlsx"'
        )
        workbook.save(response)
        return response


class InventoryPDFExportView(LoginRequiredMixin, GroupRequiredMixin, View):
    group_names = STOCK_VIEW_GROUPS

    def get(self, request, pk, *args, **kwargs):
        try:
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError:
            return HttpResponse(_("PDF export requires reportlab."), status=503)

        inventory_count = get_completed_inventory_for_export(request.user, pk)
        lines = get_inventory_result_lines(inventory_count)
        generated_at = timezone.now()
        try:
            regular_path, bold_path = self.pdf_font_paths()
        except FileNotFoundError:
            return HttpResponse(_("Unicode PDF font is not available."), status=503)
        pdfmetrics.registerFont(TTFont("InventorySans", regular_path))
        pdfmetrics.registerFont(TTFont("InventorySansBold", bold_path))

        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=14 * mm,
            leftMargin=14 * mm,
            topMargin=14 * mm,
            bottomMargin=14 * mm,
            title=f"{gettext('Інвентаризація')} {inventory_count.number}",
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "InventoryTitle",
            parent=styles["Title"],
            fontName="InventorySansBold",
            fontSize=16,
            leading=20,
            alignment=TA_CENTER,
        )
        body_style = ParagraphStyle(
            "InventoryBody",
            parent=styles["BodyText"],
            fontName="InventorySans",
            fontSize=8,
            leading=10,
        )
        header_style = ParagraphStyle(
            "InventoryHeader",
            parent=body_style,
            fontName="InventorySansBold",
        )
        section_style = ParagraphStyle(
            "InventorySection",
            parent=header_style,
            fontSize=10,
            leading=13,
            textColor=colors.HexColor("#92400E"),
        )
        result_style = ParagraphStyle(
            "InventoryResult",
            parent=body_style,
            fontSize=5.4,
            leading=6.4,
        )
        result_header_style = ParagraphStyle(
            "InventoryResultHeader",
            parent=result_style,
            fontName="InventorySansBold",
        )

        story = [
            Paragraph(
                escape(f"{gettext('Інвентаризація')} {inventory_count.number}"),
                title_style,
            ),
            Spacer(1, 5 * mm),
            Paragraph(escape(gettext("Підсумок інвентаризації")), section_style),
            Spacer(1, 2 * mm),
        ]
        summary_data = [
            [
                Paragraph(escape(str(label)), header_style),
                Paragraph(escape(self.summary_value(value)), body_style),
            ]
            for label, value in inventory_export_summary(
                inventory_count, lines, generated_at
            )
        ]
        summary_table = Table(summary_data, colWidths=[55 * mm, 105 * mm])
        summary_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d7dee8")),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f8fafc")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.extend(
            [
                summary_table,
                Spacer(1, 7 * mm),
                Paragraph(escape(gettext("Результати")), section_style),
                Spacer(1, 2 * mm),
            ]
        )

        result_data = [
            [
                Paragraph(escape(str(header)), result_header_style)
                for header in INVENTORY_RESULT_HEADERS
            ]
        ]
        for line in lines:
            row = inventory_result_row(line)
            result_data.append(
                [
                    Paragraph(escape(self.result_value(value)), result_style)
                    for value in row
                ]
            )
        result_table = Table(
            result_data,
            repeatRows=1,
            colWidths=[
                13 * mm,
                16 * mm,
                27 * mm,
                9 * mm,
                13 * mm,
                12 * mm,
                14 * mm,
                14 * mm,
                12 * mm,
                16 * mm,
                17 * mm,
                19 * mm,
            ],
        )
        result_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, 0), "InventorySansBold"),
                    ("FONTNAME", (0, 1), (-1, -1), "InventorySans"),
                    ("FONTSIZE", (0, 0), (-1, -1), 5.4),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d4ac00")),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d7dee8")),
                    ("ALIGN", (4, 1), (8, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ]
            )
        )
        story.append(result_table)
        document.build(story)

        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="inventory_{inventory_count.number}.pdf"'
        )
        return response

    @staticmethod
    def summary_value(value):
        if value is not None and hasattr(value, "tzinfo"):
            return date_format(timezone.localtime(value), "SHORT_DATETIME_FORMAT")
        return "" if value is None else str(value)

    @staticmethod
    def result_value(value):
        if value not in (None, "") and hasattr(value, "tzinfo"):
            return date_format(timezone.localtime(value), "SHORT_DATETIME_FORMAT")
        return "" if value is None else str(value)

    @staticmethod
    def pdf_font_paths():
        regular_paths = (
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        )
        bold_paths = (
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf",
            "/System/Library/Fonts/Supplemental/Arial.ttf",
        )
        regular = next((path for path in regular_paths if Path(path).exists()), None)
        bold = next((path for path in bold_paths if Path(path).exists()), regular)
        if not regular:
            raise FileNotFoundError("Unicode PDF font is not available.")
        return regular, bold


class InventoryListView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    group_names = STOCK_VIEW_GROUPS
    model = InventoryCount
    template_name = "core/inventory_list.html"
    context_object_name = "inventory_counts"
    paginate_by = 50

    def get_queryset(self):
        return InventoryCount.objects.select_related(
            "warehouse", "location", "created_by"
        ).filter(
            warehouse__in=get_accessible_warehouses(self.request.user)
        ).order_by("-started_at", "-id")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["can_edit_inventory"] = user_in_groups(self.request.user, STOCK_EDIT_GROUPS)
        return context


class InventoryCreateView(LoginRequiredMixin, GroupRequiredMixin, FormView):
    group_names = STOCK_EDIT_GROUPS
    template_name = "core/inventory_create.html"
    form_class = InventoryCountCreateForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["request_user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        inventory_count = create_inventory_count(
            warehouse=form.cleaned_data["warehouse"],
            location=form.cleaned_data["location"],
            user=self.request.user,
            comment=form.cleaned_data["comment"],
        )
        messages.success(self.request, _("Інвентаризацію створено."))
        return redirect("inventory_count", pk=inventory_count.pk)


class InventoryDetailView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    group_names = STOCK_VIEW_GROUPS
    template_name = "core/inventory_detail.html"

    def get_inventory_count(self):
        return get_object_or_404(
            InventoryCount.objects.select_related(
                "warehouse", "location", "created_by"
            ).filter(warehouse__in=get_accessible_warehouses(self.request.user)),
            pk=self.kwargs["pk"],
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventory_count = self.get_inventory_count()
        context["inventory_count"] = inventory_count
        lines = list(
            inventory_count.lines.select_related(
                "inventory_count__warehouse",
                "item",
                "item__barcode",
                "location",
                "location__warehouse",
            )
        )
        for line in lines:
            reconcile_inventory_line(line)
        variances = [line.variance_qty for line in lines]
        context["lines"] = lines
        context["inventory_summary"] = {
            "total_lines": len(lines),
            "difference_lines": sum(variance != 0 for variance in variances),
            "total_surplus": sum(
                (variance for variance in variances if variance > 0), start=0
            ),
            "total_shortage": abs(
                sum((variance for variance in variances if variance < 0), start=0)
            ),
        }
        context["can_export_inventory"] = user_in_groups(
            self.request.user, STOCK_VIEW_GROUPS
        )
        context["can_edit_inventory"] = user_in_groups(
            self.request.user, STOCK_EDIT_GROUPS
        )
        context["can_complete_inventory"] = user_in_groups(
            self.request.user, STOCK_EDIT_GROUPS
        )
        return context


class InventoryCompleteView(LoginRequiredMixin, GroupRequiredMixin, View):
    group_names = STOCK_EDIT_GROUPS

    def post(self, request, pk):
        inventory_count = get_object_or_404(
            InventoryCount.objects.filter(
                warehouse__in=get_accessible_warehouses(request.user)
            ),
            pk=pk,
        )
        try:
            complete_inventory_count(inventory_count=inventory_count, user=request.user)
        except InventoryServiceError as exc:
            messages.error(request, exc)
        else:
            messages.success(
                request,
                _("Інвентаризацію завершено. Залишки скориговано."),
            )
        return redirect("inventory_detail", pk=pk)


class InventoryCountView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    group_names = STOCK_EDIT_GROUPS
    template_name = "core/inventory_count.html"

    def get_inventory_count(self):
        return get_object_or_404(
            InventoryCount.objects.select_related(
                "warehouse", "location", "created_by"
            ).filter(warehouse__in=get_accessible_warehouses(self.request.user)),
            pk=self.kwargs["pk"],
            status=InventoryCount.Status.IN_PROGRESS,
        )

    def get_lines(self, inventory_count):
        queryset = inventory_count.lines.select_related(
            "inventory_count__warehouse",
            "item",
            "item__barcode",
            "location",
            "location__warehouse",
        )
        query = self.request.GET.get("q", "").strip()
        if query:
            queryset = queryset.filter(
                Q(item__name__icontains=query)
                | Q(item__internal_code__icontains=query)
                | Q(barcode__icontains=query)
                | Q(item__barcode__barcode__icontains=query)
                | Q(location__name__icontains=query)
                | Q(location__barcode__barcode__icontains=query)
            )
        return queryset

    def build_line_forms(self, lines, data=None):
        return [
            InventoryCountLineForm(data=data, instance=line, prefix=f"line-{line.pk}")
            for line in lines
        ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventory_count = kwargs.get("inventory_count") or self.get_inventory_count()
        lines = list(kwargs.get("lines") or self.get_lines(inventory_count))
        for line in lines:
            reconcile_inventory_line(line)
        context["inventory_count"] = inventory_count
        context["lines"] = lines
        context["line_forms"] = kwargs.get("line_forms") or self.build_line_forms(lines)
        context["search_query"] = self.request.GET.get("q", "").strip()
        return context

    def post(self, request, *args, **kwargs):
        inventory_count = self.get_inventory_count()
        lines = list(self.get_lines(inventory_count))
        submitted_barcode = request.POST.get("barcode", "").strip()
        if submitted_barcode:
            item = resolve_item_barcode(submitted_barcode)
            if item is None:
                messages.error(request, _("Товар за цим штрихкодом не знайдено."))
                return self.render_to_response(
                    self.get_context_data(inventory_count=inventory_count, lines=lines)
                )
            if not any(line.item_id == item.pk for line in lines):
                messages.error(request, _("Товар не знайдено у цій інвентаризації."))
                return self.render_to_response(
                    self.get_context_data(inventory_count=inventory_count, lines=lines)
                )
        line_forms = self.build_line_forms(lines, data=request.POST)
        if not all(form.is_valid() for form in line_forms):
            return self.render_to_response(
                self.get_context_data(
                    inventory_count=inventory_count, lines=lines, line_forms=line_forms
                )
            )
        for form in line_forms:
            actual_qty = form.cleaned_data.get("actual_qty")
            if actual_qty is None:
                continue
            update_inventory_line_actual_qty(
                line=form.instance,
                actual_qty=actual_qty,
                user=request.user,
                comment=form.cleaned_data.get("comment", ""),
            )
        messages.success(request, _("Підрахунок збережено."))
        url = reverse("inventory_count", kwargs={"pk": inventory_count.pk})
        query = request.GET.urlencode()
        if query:
            url = f"{url}?{query}"
        return redirect(url)

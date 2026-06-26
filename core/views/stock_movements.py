from collections import Counter

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.utils import timezone
from django.utils.translation import gettext, gettext_lazy as _
from django.views.generic import FormView, ListView, TemplateView, View

from ..forms import (
    StockMovementCancellationForm,
    StockMovementFilterForm,
    StockOperationAuditFilterForm,
)
from ..models import StockBalance, StockMovement
from ..permissions import (
    STOCK_EDIT_GROUPS,
    STOCK_VIEW_GROUPS,
    MANAGEMENT_GROUPS,
    GroupRequiredMixin,
    can_cancel_movement,
)
from ..services.filter_memory import apply_remembered_filters, build_redirect_url, querydict_from_params
from ..services.analytics.data_quality import filter_quality_check
from ..services.movements import trusted_business_movements
from ..services.stock import StockServiceError
from ..services.stock_cancellation import cancel_stock_movement
from ..services.warehouse_access import restrict_stock_movement_queryset_for_user
from .stock_operations import SelfServiceShellContextMixin


def get_stock_operation_audit_filter_form(user, params):
    return StockOperationAuditFilterForm(params or None, request_user=user)


def get_stock_operation_audit_queryset(user, params):
    queryset = restrict_stock_movement_queryset_for_user(
        user,
        StockMovement.objects.select_related(
            "item",
            "item__barcode",
            "item__unit",
            "source_location__warehouse",
            "source_warehouse",
            "destination_location__warehouse",
            "destination_warehouse",
            "recipient",
            "created_by",
            "performed_by",
            "cancelled_by",
            "inventory_count",
            "reversal_of",
            "cancellation_movement",
        ),
    )
    form = get_stock_operation_audit_filter_form(user, params)
    if form.is_valid():
        cd = form.cleaned_data
        if cd.get("date_from"):
            queryset = queryset.filter(created_at__date__gte=cd["date_from"])
        if cd.get("date_to"):
            queryset = queryset.filter(created_at__date__lte=cd["date_to"])
        if cd.get("movement_type"):
            queryset = queryset.filter(movement_type=cd["movement_type"])
        if cd.get("warehouse"):
            warehouse = cd["warehouse"]
            queryset = queryset.filter(
                Q(source_warehouse=warehouse)
                | Q(destination_warehouse=warehouse)
                | Q(source_location__warehouse=warehouse)
                | Q(destination_location__warehouse=warehouse)
            )
        if cd.get("location"):
            location = cd["location"]
            queryset = queryset.filter(
                Q(source_location=location) | Q(destination_location=location)
            )
        if cd.get("q"):
            query = cd["q"]
            queryset = queryset.filter(
                Q(item__name__icontains=query)
                | Q(item__internal_code__icontains=query)
                | Q(item__barcode__barcode__icontains=query)
            )
        if cd.get("quantity_from") is not None:
            queryset = queryset.filter(qty__gte=cd["quantity_from"])
        if cd.get("quantity_to") is not None:
            queryset = queryset.filter(qty__lte=cd["quantity_to"])
        if cd.get("recipient"):
            queryset = queryset.filter(recipient=cd["recipient"])
        if cd.get("document"):
            document = cd["document"]
            queryset = queryset.filter(
                Q(document_number__icontains=document)
                | Q(comment__icontains=document)
                | Q(cancellation_reason__icontains=document)
            )
        if cd.get("user"):
            audit_user = cd["user"]
            queryset = queryset.filter(
                Q(created_by=audit_user)
                | Q(performed_by=audit_user)
                | Q(cancelled_by=audit_user)
            )
        if cd.get("cancelled") == "yes":
            queryset = queryset.filter(is_cancelled=True)
        elif cd.get("cancelled") == "no":
            queryset = queryset.filter(is_cancelled=False)
        if cd.get("inventory_related") == "yes":
            queryset = queryset.filter(inventory_count__isnull=False)
        elif cd.get("inventory_related") == "no":
            queryset = queryset.filter(inventory_count__isnull=True)
    return queryset.order_by("-created_at", "-id")


def audit_user_display_name(user):
    if not user:
        return ""
    return (user.get_full_name() or user.get_username()).strip()


def local_naive(value):
    if not value:
        return None
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.replace(tzinfo=None)


class StockMovementListView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    group_names = STOCK_VIEW_GROUPS
    model = StockMovement
    template_name = "core/stockmovement_list.html"
    context_object_name = "movements"
    paginate_by = 50
    page_key = "movement_list"
    page_title = _("Журнал операцій")
    page_action_label = _("Нова операція")
    reset_url_name = "movement_list"
    is_document_registry = False

    def get(self, request, *args, **kwargs):
        params, used_remembered_filters, should_redirect = apply_remembered_filters(request, self.page_key)
        self.used_remembered_filters = used_remembered_filters
        if should_redirect:
            return redirect(build_redirect_url(request.path, params))
        self.effective_get = querydict_from_params(params) if params else request.GET
        return super().get(request, *args, **kwargs)

    def get_filter_form(self):
        return StockMovementFilterForm(
            getattr(self, "effective_get", self.request.GET) or None,
            request_user=self.request.user,
        )

    def get_queryset(self):
        queryset = restrict_stock_movement_queryset_for_user(
            self.request.user,
            StockMovement.objects.select_related(
                "item",
                "item__barcode",
                "source_location",
                "source_warehouse",
                "source_location__warehouse",
                "destination_location",
                "destination_warehouse",
                "destination_location__warehouse",
                "recipient",
                "inventory_count",
                "performed_by",
                "cancelled_by",
                "reversal_of",
            ),
        )
        form = self.get_filter_form()
        if not form.is_valid():
            return queryset
        cd = form.cleaned_data
        if cd.get("report_scope") == "business":
            queryset = trusted_business_movements(queryset)
        if cd.get("movement_type"):
            queryset = queryset.filter(movement_type=cd["movement_type"])
        if cd.get("item"):
            queryset = queryset.filter(item=cd["item"])
        if cd.get("item_id"):
            queryset = queryset.filter(item_id=cd["item_id"])
        if cd.get("recipient"):
            queryset = queryset.filter(recipient=cd["recipient"])
        if cd.get("recipient_id"):
            queryset = queryset.filter(recipient_id=cd["recipient_id"])
        if cd.get("warehouse"):
            queryset = queryset.filter(
                Q(source_warehouse=cd["warehouse"]) | Q(destination_warehouse=cd["warehouse"]) | Q(source_location__warehouse=cd["warehouse"]) | Q(destination_location__warehouse=cd["warehouse"])
            )
        if cd.get("location"):
            queryset = queryset.filter(
                Q(source_location=cd["location"]) | Q(destination_location=cd["location"])
            )
        if cd.get("date_from"):
            queryset = queryset.filter(occurred_at__date__gte=cd["date_from"])
        if cd.get("date_to"):
            queryset = queryset.filter(occurred_at__date__lte=cd["date_to"])
        if cd.get("issue_reason"):
            queryset = queryset.filter(issue_reason=cd["issue_reason"])
        if cd.get("department"):
            queryset = queryset.filter(department__icontains=cd["department"])
        if cd.get("usage_place_id"):
            queryset = queryset.filter(department__iexact=cd["usage_place_id"])
        if cd.get("document_number"):
            queryset = queryset.filter(document_number__icontains=cd["document_number"])
        if cd.get("no_document"):
            queryset = queryset.filter(Q(document_number__isnull=True) | Q(document_number=""))
        if cd.get("missing_recipient"):
            queryset = queryset.filter(recipient__isnull=True)
        if cd.get("missing_usage_place"):
            queryset = queryset.filter(Q(department__isnull=True) | Q(department=""))
        if cd.get("missing_destination"):
            queryset = queryset.filter(destination_location__isnull=True)
        if cd.get("invalid_qty"):
            queryset = queryset.filter(qty__lte=0)
        if cd.get("q"):
            q = cd["q"]
            queryset = queryset.filter(
                Q(item__name__icontains=q) | Q(item__internal_code__icontains=q) | Q(item__barcode__barcode__icontains=q)
            )
        if cd.get("quality_check"):
            queryset = filter_quality_check(queryset, cd["quality_check"])
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["filter_form"] = self.get_filter_form()
        context["used_remembered_filters"] = getattr(self, "used_remembered_filters", False)
        context["page_title"] = self.page_title
        context["page_action_label"] = self.page_action_label
        context["reset_url"] = reverse(self.reset_url_name)
        context["is_document_registry"] = self.is_document_registry
        context["business_report_scope"] = (
            context["filter_form"].is_valid()
            and context["filter_form"].cleaned_data.get("report_scope") == "business"
        )
        return context


class StockDocumentListView(StockMovementListView):
    page_key = "stock_documents"
    page_title = _("Реєстр документів")
    page_action_label = _("Нова операція")
    reset_url_name = "stock_documents"
    is_document_registry = True


class StockOperationAuditView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    group_names = MANAGEMENT_GROUPS
    model = StockMovement
    template_name = "core/management/stock_operation_audit.html"
    context_object_name = "movements"
    paginate_by = 50

    def get_filter_form(self):
        return get_stock_operation_audit_filter_form(self.request.user, self.request.GET)

    def get_queryset(self):
        return get_stock_operation_audit_queryset(self.request.user, self.request.GET)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query_params = self.request.GET.copy()
        query_params.pop("page", None)
        filter_form = self.get_filter_form()
        context["filter_form"] = filter_form
        context["active_filter_names"] = [
            name
            for name in filter_form.fields
            if self.request.GET.get(name) not in (None, "")
        ]
        context["filter_query"] = query_params.urlencode()
        return context


class StockOperationAuditXLSXExportView(LoginRequiredMixin, GroupRequiredMixin, View):
    group_names = MANAGEMENT_GROUPS

    headers = [
        _("Дата й час операції"),
        _("Дата й час створення"),
        _("Тип операції"),
        _("Статус"),
        _("Код товару"),
        _("Назва товару"),
        _("Кількість"),
        _("Одиниця"),
        _("Склад-відправник"),
        _("Склад-отримувач"),
        _("Локація-відправник"),
        _("Локація-отримувач"),
        _("Отримувач"),
        _("Документ"),
        _("Коментар / причина"),
        _("Створив"),
        _("Анулювано"),
        _("Анулював"),
        _("Час анулювання"),
        _("Рух анулювання"),
        _("Зворотний рух"),
        _("Інвентаризація"),
    ]

    def get(self, request, *args, **kwargs):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return HttpResponse(_("XLSX export requires openpyxl."), status=503)

        movements = list(get_stock_operation_audit_queryset(request.user, request.GET))
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        report = workbook.create_sheet("Audit report")

        self.populate_summary(summary, movements, request)
        report.append([str(header) for header in self.headers])
        for movement in movements:
            report.append(self.movement_row(movement))

        header_fill = PatternFill("solid", fgColor="D4AC00")
        cancelled_fill = PatternFill("solid", fgColor="FDECEC")
        for cell in report[1]:
            cell.font = Font(bold=True, color="101828")
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row_number, movement in enumerate(movements, start=2):
            if movement.is_cancelled:
                for cell in report[row_number]:
                    cell.fill = cancelled_fill
            for column in (1, 2, 19):
                report.cell(row=row_number, column=column).number_format = "yyyy-mm-dd hh:mm:ss"
            report.cell(row=row_number, column=7).number_format = "#,##0.000"

        widths = [
            21, 21, 19, 16, 18, 32, 14, 14, 25, 25, 25,
            25, 24, 20, 38, 24, 14, 24, 21, 20, 20, 22,
        ]
        for index, width in enumerate(widths, start=1):
            report.column_dimensions[report.cell(row=1, column=index).column_letter].width = width
        report.freeze_panes = "A2"
        report.auto_filter.ref = report.dimensions

        for row in summary.iter_rows():
            row[0].font = Font(bold=True, color="92400E")
        summary["A1"].font = Font(bold=True, size=14, color="92400E")
        summary["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
        summary.column_dimensions["A"].width = 34
        summary.column_dimensions["B"].width = 44

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = timezone.localtime().strftime("stock-operation-audit-%Y%m%d-%H%M.xlsx")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    def populate_summary(self, sheet, movements, request):
        form = get_stock_operation_audit_filter_form(request.user, request.GET)
        cleaned_data = form.cleaned_data if form.is_bound and form.is_valid() else {}
        active_filters = []
        for name, field in form.fields.items():
            value = cleaned_data.get(name)
            if value not in (None, ""):
                active_filters.append(f"{field.label}: {self.filter_value(field, value)}")

        sheet.append([gettext("Звіт аудиту складських операцій")])
        sheet.append([gettext("Згенеровано"), local_naive(timezone.now())])
        sheet.append([gettext("Згенерував"), audit_user_display_name(request.user)])
        sheet.append(
            [
                gettext("Активні фільтри"),
                "; ".join(active_filters) or gettext("Активні фільтри відсутні"),
            ]
        )
        sheet.append([gettext("Усього рухів"), len(movements)])
        sheet.append(
            [gettext("Усього анульованих рухів"), sum(m.is_cancelled for m in movements)]
        )
        sheet.append(
            [
                gettext("Усього рухів з інвентаризації"),
                sum(bool(m.inventory_count_id) for m in movements),
            ]
        )
        sheet.append([])
        sheet.append([gettext("Підсумки за типом операції"), gettext("Кількість")])
        for label, count in Counter(m.get_movement_type_display() for m in movements).most_common():
            sheet.append([label, count])
        sheet.append([])
        sheet.append([gettext("Підсумки за складом"), gettext("Кількість")])
        warehouses = Counter()
        for movement in movements:
            for warehouse in {
                movement.resolved_source_warehouse,
                movement.resolved_destination_warehouse,
            }:
                if warehouse:
                    warehouses[str(warehouse)] += 1
        for warehouse, count in warehouses.most_common():
            sheet.append([warehouse, count])

    @staticmethod
    def filter_value(field, value):
        if getattr(field, "choices", None):
            return dict(field.choices).get(value, value)
        return str(value)

    @staticmethod
    def movement_row(movement):
        created_by = movement.created_by or movement.performed_by
        comment_reason = " / ".join(
            value for value in (movement.comment, movement.cancellation_reason) if value
        )
        return [
            local_naive(movement.occurred_at),
            local_naive(movement.created_at),
            movement.get_movement_type_display(),
            gettext("Анулювано") if movement.is_cancelled else gettext("Активно"),
            movement.item.internal_code,
            movement.item.name,
            float(movement.qty),
            movement.item.unit.symbol,
            str(movement.resolved_source_warehouse or ""),
            str(movement.resolved_destination_warehouse or ""),
            str(movement.source_location or ""),
            str(movement.destination_location or ""),
            str(movement.recipient or ""),
            movement.document_number,
            comment_reason,
            audit_user_display_name(created_by),
            gettext("Так") if movement.is_cancelled else gettext("Ні"),
            audit_user_display_name(movement.cancelled_by),
            local_naive(movement.cancelled_at),
            f"#{movement.cancellation_movement_id}" if movement.cancellation_movement_id else "",
            f"#{movement.reversal_of_id}" if movement.reversal_of_id else "",
            movement.inventory_count.number if movement.inventory_count_id else "",
        ]


class StockReceiveResultView(
    LoginRequiredMixin, SelfServiceShellContextMixin, GroupRequiredMixin, TemplateView
):
    group_names = STOCK_EDIT_GROUPS
    template_name = "core/stock_receive_result.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["movement"] = get_object_or_404(
            restrict_stock_movement_queryset_for_user(
                self.request.user,
                StockMovement.objects.select_related(
                    "item",
                    "item__barcode",
                    "destination_location",
                    "destination_warehouse",
                "destination_location__warehouse",
                    "recipient",
                    "performed_by",
                    "cancelled_by",
                    "reversal_of",
                ),
            ),
            pk=self.kwargs["pk"],
            movement_type__in=[StockMovement.MovementType.IN, StockMovement.MovementType.RETURN],
        )
        return context


class StockMovementCancelView(LoginRequiredMixin, FormView):
    template_name = "core/stock_movement_cancel.html"
    form_class = StockMovementCancellationForm

    def dispatch(self, request, *args, **kwargs):
        self.movement = get_object_or_404(
            StockMovement.objects.select_related(
                "item",
                "source_location",
                "source_warehouse",
                "source_location__warehouse",
                "destination_location",
                "destination_warehouse",
                "destination_location__warehouse",
                "recipient",
                "performed_by",
                "cancelled_by",
            ),
            pk=kwargs["pk"],
        )
        if not can_cancel_movement(request.user, self.movement):
            raise PermissionDenied(_("У вас немає прав для перегляду цієї сторінки."))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["movement"] = self.movement
        return context

    def form_valid(self, form):
        try:
            cancellation_movement = cancel_stock_movement(
                movement=self.movement,
                cancelled_by=self.request.user,
                reason=form.cleaned_data["reason"],
                request=self.request,
            )
        except StockServiceError as exc:
            message = str(exc)
            messages.error(self.request, message)
            form.add_error(None, message)
            return self.form_invalid(form)
        messages.success(self.request, _("Рух анульовано."))
        return redirect("stock_movement_print", pk=cancellation_movement.pk)


class StockMovementPrintView(
    LoginRequiredMixin, SelfServiceShellContextMixin, GroupRequiredMixin, TemplateView
):
    group_names = STOCK_VIEW_GROUPS
    template_name = "core/stock_movement_print.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        movement = get_object_or_404(
            restrict_stock_movement_queryset_for_user(
                self.request.user,
                StockMovement.objects.select_related(
                    "item",
                    "item__barcode",
                    "source_location",
                    "source_warehouse",
                "source_location__warehouse",
                    "destination_location",
                    "destination_warehouse",
                "destination_location__warehouse",
                    "recipient",
                    "performed_by",
                    "cancelled_by",
                    "reversal_of",
                ),
            ),
            pk=self.kwargs["pk"],
        )
        location = movement.source_location or movement.destination_location
        is_self_service_movement = movement.movement_type in {
            StockMovement.MovementType.OUT,
            StockMovement.MovementType.RETURN,
        }
        if movement.movement_type == StockMovement.MovementType.OUT:
            operation_type = _("Видача товару")
            responsible_label = _("Хто взяв товар")
        elif movement.movement_type == StockMovement.MovementType.RETURN:
            operation_type = _("Повернення товару")
            responsible_label = _("Хто повернув товар")
        elif movement.movement_type == StockMovement.MovementType.IN:
            operation_type = _("Прихід товару")
            responsible_label = _("Отримувач / відповідальний")
        else:
            operation_type = movement.get_movement_type_display()
            responsible_label = _("Отримувач / відповідальний")
        labels = {
            "title": _("Контрольний талон складської операції"),
            "print": _("Друк"),
            "back": _("Назад"),
            "operation_id": _("ID операції"),
            "operation_type": _("Тип операції"),
            "occurred_at": _("Дата і час операції"),
            "item": _("Товар"),
            "internal_code": _("Внутрішній код"),
            "barcode": _("Штрихкод"),
            "qty": _("Кількість"),
            "warehouse": _("Склад"),
            "location": _("Локація"),
            "responsible": responsible_label,
            "department": _("Цех / місце використання"),
            "comment_document": _("Коментар / документ"),
            "video_time": _("Час для перевірки по відео:"),
        }
        context.update(
            {
                "labels": labels,
                "movement": movement,
                "operation_type": operation_type,
                "location": location,
                "warehouse": movement.resolved_source_warehouse or movement.resolved_destination_warehouse,
                "is_self_service_movement": is_self_service_movement,
            }
        )
        return context


class StockIssueResultView(
    LoginRequiredMixin, SelfServiceShellContextMixin, GroupRequiredMixin, TemplateView
):
    group_names = STOCK_EDIT_GROUPS
    template_name = "core/stock_issue_result.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        movement = get_object_or_404(
            restrict_stock_movement_queryset_for_user(
                self.request.user,
                StockMovement.objects.select_related(
                    "item",
                    "source_location",
                    "source_warehouse",
                "source_location__warehouse",
                    "recipient",
                    "performed_by",
                    "cancelled_by",
                    "reversal_of",
                ),
            ),
            pk=self.kwargs["pk"],
            movement_type=StockMovement.MovementType.OUT,
        )
        context["movement"] = movement
        context["balance_after"] = get_object_or_404(
            StockBalance, item=movement.item, warehouse=movement.resolved_source_warehouse, is_active=True
        ).qty
        return context


class StockWriteOffResultView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    group_names = STOCK_VIEW_GROUPS
    template_name = "core/stock_writeoff_result.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["movement"] = get_object_or_404(
            restrict_stock_movement_queryset_for_user(
                self.request.user,
                StockMovement.objects.select_related(
                    "item", "source_location", "source_warehouse",
                "source_location__warehouse", "performed_by"
                ),
            ),
            pk=self.kwargs["pk"],
            movement_type=StockMovement.MovementType.WRITEOFF,
        )
        return context


class StockTransferResultView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    group_names = STOCK_VIEW_GROUPS
    template_name = "core/stock_transfer_result.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["movement"] = get_object_or_404(
            restrict_stock_movement_queryset_for_user(
                self.request.user,
                StockMovement.objects.select_related(
                    "item",
                    "source_location",
                    "source_warehouse",
                "source_location__warehouse",
                    "destination_location",
                    "destination_warehouse",
                "destination_location__warehouse",
                    "performed_by",
                    "cancelled_by",
                    "reversal_of",
                ),
            ),
            pk=self.kwargs["pk"],
            movement_type=StockMovement.MovementType.TRANSFER,
        )
        return context

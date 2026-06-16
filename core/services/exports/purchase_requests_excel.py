from django.utils.timezone import localtime
from django.utils.translation import gettext_lazy as _

from core.models import PurchaseRequest


def _display_user(user):
    if not user:
        return ""
    return user.get_full_name() or user.get_username()


def _local_datetime(value):
    if not value:
        return ""
    return localtime(value).replace(tzinfo=None)


STATUS_FILLS = {
    "green": "D9EAD3",
    "yellow": "FFF2CC",
    "red": "F4CCCC",
}


def _status_fill(value):
    positive = {
        PurchaseRequest.ApprovalStatus.APPROVED,
        PurchaseRequest.PaymentStatus.PAID,
        PurchaseRequest.PaymentStatus.NOT_REQUIRED,
        PurchaseRequest.DeliveryStatus.DELIVERED,
    }
    pending = {
        PurchaseRequest.ApprovalStatus.PENDING,
        PurchaseRequest.PaymentStatus.INVOICE_NOT_RECEIVED,
        PurchaseRequest.PaymentStatus.INVOICE_RECEIVED,
        PurchaseRequest.PaymentStatus.SENT_FOR_PAYMENT,
        PurchaseRequest.DeliveryStatus.NOT_SHIPPED,
        PurchaseRequest.DeliveryStatus.IN_TRANSIT,
    }
    problem = {
        PurchaseRequest.ApprovalStatus.REJECTED,
        PurchaseRequest.DeliveryStatus.CANCELLED,
    }
    if value in positive:
        return STATUS_FILLS["green"]
    if value in pending:
        return STATUS_FILLS["yellow"]
    if value in problem:
        return STATUS_FILLS["red"]
    return None


def build_purchase_requests_workbook(purchase_requests):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    headers = [
        _("Дата створення"),
        _("Назва товару"),
        _("Опис потреби"),
        _("Посилання на товар"),
        _("Кількість"),
        _("Одиниця виміру"),
        _("Вартість за одиницю, грн"),
        _("Сума, грн"),
        _("Тип замовлення"),
        _("Статус погодження"),
        _("Статус оплати"),
        _("Статус доставки"),
        _("Заявник"),
        _("Ким погоджено"),
        _("Дата погодження"),
        _("Ким відхилено"),
        _("Дата відхилення"),
        _("Коментар відхилення"),
    ]
    widths = [19, 34, 42, 42, 12, 16, 22, 16, 18, 20, 24, 20, 24, 24, 21, 24, 21, 34]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Purchase requests"
    sheet.append([str(header) for header in headers])

    status_values_by_row = {}
    for purchase_request in purchase_requests:
        row_number = sheet.max_row + 1
        sheet.append(
            [
                _local_datetime(purchase_request.created_at),
                purchase_request.title,
                purchase_request.need_description,
                purchase_request.product_url,
                purchase_request.requested_qty,
                purchase_request.unit,
                purchase_request.unit_price_uah or "",
                purchase_request.total_price_uah or "",
                purchase_request.get_order_type_display(),
                purchase_request.get_approval_status_display(),
                purchase_request.get_payment_status_display(),
                purchase_request.get_delivery_status_display(),
                _display_user(purchase_request.requested_by),
                _display_user(purchase_request.approved_by),
                _local_datetime(purchase_request.approved_at),
                _display_user(purchase_request.rejected_by),
                _local_datetime(purchase_request.rejected_at),
                purchase_request.rejection_comment,
            ]
        )
        status_values_by_row[row_number] = (
            purchase_request.approval_status,
            purchase_request.payment_status,
            purchase_request.delivery_status,
        )

    header_fill = PatternFill("solid", fgColor="D4AC00")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="101828")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for index, width in enumerate(widths, start=1):
        column_letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column_letter].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row in sheet.iter_rows(min_row=2):
        row[0].number_format = "yyyy-mm-dd hh:mm:ss"
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[4].number_format = "#,##0.###"
        row[6].number_format = "#,##0.00"
        row[7].number_format = "#,##0.00"
        status_values = status_values_by_row.get(row[0].row, ("", "", ""))
        for index, status_value in zip((9, 10, 11), status_values, strict=True):
            fill_color = _status_fill(status_value)
            if fill_color:
                row[index].fill = PatternFill("solid", fgColor=fill_color)
        row[13].number_format = "yyyy-mm-dd hh:mm:ss"
        row[15].number_format = "yyyy-mm-dd hh:mm:ss"
        row[17].alignment = Alignment(wrap_text=True, vertical="top")

    return workbook
